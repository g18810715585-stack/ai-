from __future__ import annotations

import json
import os
import socket
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .io_utils import write_json, write_text
from .models import Manifest, Patch, PatchOperation, SchemaBundle, SourceRef

DEFAULT_AI_TIMEOUT_SECONDS = 240

AI_PROVIDERS = {
    "chatgpt": {
        "label": "ChatGPT",
        "api_key_env": "BASEAI_API_KEY",
        "base_url_env": "BASEAI_BASE_URL",
        "model_env": "CHATGPT_MODEL",
        "default_base_url": "https://baseai.rivergame.net/v1",
        "default_model": "gpt-5.5",
        "extra_body": {},
        "temperature": 0.1,
    },
    "gemini": {
        "label": "Gemini",
        "api_key_env": "BASEAI_API_KEY",
        "base_url_env": "BASEAI_BASE_URL",
        "model_env": "GEMINI_MODEL",
        "default_base_url": "https://baseai.rivergame.net/v1",
        "default_model": "gemini-3.1-pro-preview",
        "extra_body": {},
        "temperature": 0.1,
    },
    "claude": {
        "label": "Claude",
        "api_key_env": "BASEAI_API_KEY",
        "base_url_env": "BASEAI_BASE_URL",
        "model_env": "CLAUDE_MODEL",
        "default_base_url": "https://baseai.rivergame.net/v1",
        "default_model": "claude-opus-4-8",
        "extra_body": {},
        "temperature": None,
    },
    "deepseek_v4_pro": {
        "label": "DeepSeek",
        "api_key_env": "BASEAI_API_KEY",
        "base_url_env": "BASEAI_BASE_URL",
        "model_env": "DEEPSEEK_MODEL",
        "default_base_url": "https://baseai.rivergame.net/v1",
        "default_model": "deepseek-v4-pro",
        "extra_body": {},
        "temperature": 0.1,
    },
}

PROVIDER_ALIASES = {
    "baseai": "chatgpt",
    "base_ai": "chatgpt",
    "company_bi": "chatgpt",
    "openai": "chatgpt",
    "chatgpt": "chatgpt",
    "gpt": "chatgpt",
    "gemini": "gemini",
    "google": "gemini",
    "claude": "claude",
    "anthropic": "claude",
    "deepseek": "deepseek_v4_pro",
    "deepseekv4pro": "deepseek_v4_pro",
    "deepseek_v4": "deepseek_v4_pro",
    "deepseek_v4_pro": "deepseek_v4_pro",
    "deepseek-v4-pro": "deepseek_v4_pro",
}


def _coerce(value: Any, field_type: str) -> Any:
    if value in (None, ""):
        return None
    if field_type == "int":
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    if field_type == "float":
        try:
            return float(value)
        except (TypeError, ValueError):
            return value
    if field_type == "str":
        return str(value)
    return value


def _existing_keys(path: str, sheet_name: str | None, primary_key: list[str]) -> set[tuple[Any, ...]]:
    if not path or not primary_key:
        return set()
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    indexes = {str(header): idx + 1 for idx, header in enumerate(headers) if header}
    keys: set[tuple[Any, ...]] = set()
    for row_idx in range(2, sheet.max_row + 1):
        key = tuple(sheet.cell(row_idx, indexes[field]).value for field in primary_key if field in indexes)
        if len(key) == len(primary_key) and any(item is not None for item in key):
            keys.add(key)
    return keys


def make_stub_patch(manifest: Manifest, schema: SchemaBundle, context: dict[str, Any], base_dir: str) -> Patch:
    operations: list[PatchOperation] = []
    patch_id = datetime.now(timezone.utc).strftime("patch_%Y%m%d_%H%M%S")
    config_keys: dict[str, set[tuple[Any, ...]]] = {}
    for table_name, table in schema.tables.items():
        ref = manifest.config_tables.get(table_name)
        if ref:
            path = ref.path
            if not os.path.isabs(path):
                path = os.path.abspath(os.path.join(base_dir, path))
            config_keys[table_name] = _existing_keys(path, ref.sheet or table.sheet, table.primary_key)
        else:
            config_keys[table_name] = set()

    for workbook in context.get("workbooks", []):
        for sheet in workbook.get("sheets", []):
            headers = sheet.get("headers") or []
            header_to_table_field: dict[str, tuple[str, str]] = {}
            for table_name, table in schema.tables.items():
                for header in headers:
                    if header in table.field_aliases:
                        header_to_table_field[header] = (table_name, table.field_aliases[header])
                    elif header in table.fields:
                        header_to_table_field[header] = (table_name, header)
            for source_row in sheet.get("sample_rows", []):
                rows_by_table: dict[str, dict[str, Any]] = {}
                for header, value in source_row.items():
                    if header.startswith("__") or header not in header_to_table_field:
                        continue
                    table_name, field = header_to_table_field[header]
                    table = schema.tables[table_name]
                    spec = table.fields.get(field)
                    rows_by_table.setdefault(table_name, {})[field] = _coerce(value, spec.type if spec else "any")
                for table_name, row in rows_by_table.items():
                    table = schema.tables[table_name]
                    if not table.primary_key or not all(row.get(field) not in (None, "") for field in table.primary_key):
                        continue
                    for field, spec in table.fields.items():
                        if row.get(field) in (None, "") and spec.default is not None:
                            row[field] = spec.default
                    key = tuple(row.get(field) for field in table.primary_key)
                    exists = key in config_keys.get(table_name, set())
                    safe_set = {
                        field: value
                        for field, value in row.items()
                        if field not in table.block_update_fields and (not table.allow_update_fields or field in table.allow_update_fields)
                    }
                    confidence = 0.86 if exists else 0.82
                    if exists:
                        operations.append(
                            PatchOperation(
                                op="update",
                                target_table=table_name,
                                match={field: row[field] for field in table.primary_key},
                                set=safe_set,
                                source_ref=SourceRef(workbook=workbook.get("source_id"), sheet=sheet.get("name"), row=source_row.get("__row")),
                                reason="stub draft matched an existing primary key and generated a supervised field-level update",
                                confidence=confidence,
                                risk_level="medium",
                                needs_confirmation=True,
                            )
                        )
                    else:
                        operations.append(
                            PatchOperation(
                                op="insert",
                                target_table=table_name,
                                rows=[row],
                                source_ref=SourceRef(workbook=workbook.get("source_id"), sheet=sheet.get("name"), row=source_row.get("__row")),
                                reason="stub draft did not find the primary key in the target table and generated an insert",
                                confidence=confidence,
                                risk_level="medium",
                                needs_confirmation=True,
                            )
                        )
    return Patch(patch_id=patch_id, project=manifest.project, mode=manifest.mode, operations=operations)


def normalize_ai_provider(provider: str | None) -> str:
    key = str(provider or "chatgpt").strip().lower().replace("-", "_")
    return PROVIDER_ALIASES.get(key, "chatgpt")


def _resolve_ai_runtime(manifest: Manifest) -> dict[str, Any]:
    provider_id = normalize_ai_provider(manifest.ai.provider)
    provider = dict(AI_PROVIDERS[provider_id])
    return provider | {"id": provider_id}


def call_baseai(manifest: Manifest, context: dict[str, Any], raw_response_path: Path | None = None) -> Patch:
    runtime = _resolve_ai_runtime(manifest)
    api_key = os.environ.get(runtime["api_key_env"])
    if not api_key:
        raise RuntimeError(f"缺少真实 AI Key：请在项目根目录 .env 里配置 {runtime['api_key_env']}=你的Key，或切回本地草案模式。")
    base_url = os.environ.get(runtime["base_url_env"], runtime["default_base_url"]).rstrip("/")
    model = os.environ.get(runtime["model_env"], runtime["default_model"])
    url = f"{base_url}/chat/completions"
    body = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are an AI meta configuration agent. Return only one strict JSON object and no markdown. "
                    "The JSON must match this Patch shape: patch_id, project, mode, operations, generated_by. "
                    "Each operation must include op, target_table, source_ref, reason, confidence, risk_level, "
                    "needs_confirmation, and the required match/set/rows fields for the op. "
                    "Allowed op values are exactly insert, update, delete_where, replace_group; use insert with rows "
                    "for one or many new rows, never insert_rows and never insert+set. "
                    "Use only target tables and fields from schema. Treat run_instruction as the highest-priority "
                    "temporary instruction for this run, but do not convert it into long-term memory. When planning "
                    "rows, activity templates, field dictionary, matched experience, similar cases, structured "
                    "corrections, field mappings, or relationship_map provide evidence, generate a supervised patch for the "
                    "evidenced fields and mark uncertain or high-risk operations as needs_confirmation=true. "
                    "Do not return an empty patch only because some recommended fields are missing; omit uncertain "
                    "fields or use high risk. Return operations: [] only when there is no writable target table, "
                    "no usable primary key/insert-row evidence, or no schema field can be mapped safely."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(context, ensure_ascii=False),
            },
        ],
        "response_format": {"type": "json_object"},
        **runtime["extra_body"],
    }
    if runtime["temperature"] is not None:
        body["temperature"] = runtime["temperature"]
    request = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    timeout = int(os.environ.get("AI_REQUEST_TIMEOUT_SECONDS", DEFAULT_AI_TIMEOUT_SECONDS))
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except (TimeoutError, socket.timeout) as exc:
        raise RuntimeError(
            f"{runtime['label']} 请求超时：等待 {timeout} 秒仍未返回。工具已压缩上下文；如果仍失败，可换更快模型或设置 AI_REQUEST_TIMEOUT_SECONDS=360。"
        ) from exc
    except urllib.error.HTTPError as exc:
        message = exc.read().decode("utf-8", errors="replace")[:1000]
        raise RuntimeError(f"{runtime['label']} 请求失败：HTTP {exc.code} {message}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"{runtime['label']} 网络连接失败：{exc.reason}") from exc
    if raw_response_path:
        write_json(raw_response_path, {"provider": runtime["id"], "base_url": base_url, "model": model, "response": payload})
    content = payload["choices"][0]["message"]["content"]
    try:
        parsed = json.loads(content)
    except json.JSONDecodeError as exc:
        if raw_response_path:
            write_text(raw_response_path.with_name("ai-invalid-content.txt"), content)
        raise RuntimeError(f"真实 AI 返回的内容不是合法 JSON：{exc}") from exc
    try:
        return Patch.model_validate(parsed)
    except Exception as exc:
        if raw_response_path:
            write_json(raw_response_path.with_name("ai-invalid-patch.json"), parsed)
        raise RuntimeError(f"真实 AI 返回的 patch 不符合 Schema：{exc}") from exc


def call_relationship_ai(manifest: Manifest, relationship_context: dict[str, Any], raw_response_path: Path | None = None) -> dict[str, Any]:
    runtime = _resolve_ai_runtime(manifest)
    api_key = os.environ.get(runtime["api_key_env"])
    if not api_key:
        raise RuntimeError(f"缺少真实 AI Key：请在项目根目录 .env 里配置 {runtime['api_key_env']}=你的Key，或关闭 AI 解释。")
    base_url = os.environ.get(runtime["base_url_env"], runtime["default_base_url"]).rstrip("/")
    model = os.environ.get(runtime["model_env"], runtime["default_model"])
    url = f"{base_url}/chat/completions"
    body = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You explain game meta table relationships. Return only strict JSON with keys: "
                    "summary, recommended_tables, relation_notes, needs_confirmation. "
                    "Do not invent relations; only explain or rank relations present in the provided relationship map."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(relationship_context, ensure_ascii=False),
            },
        ],
        "response_format": {"type": "json_object"},
        **runtime["extra_body"],
    }
    if runtime["temperature"] is not None:
        body["temperature"] = runtime["temperature"]
    request = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        message = exc.read().decode("utf-8", errors="replace")[:1000]
        raise RuntimeError(f"{runtime['label']} 关系解释请求失败：HTTP {exc.code} {message}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"{runtime['label']} 关系解释网络连接失败：{exc.reason}") from exc
    if raw_response_path:
        write_json(raw_response_path, {"provider": runtime["id"], "base_url": base_url, "model": model, "response": payload})
    content = payload["choices"][0]["message"]["content"]
    try:
        parsed = json.loads(content)
    except json.JSONDecodeError as exc:
        if raw_response_path:
            write_text(raw_response_path.with_name("relationship-ai-invalid-content.txt"), content)
        raise RuntimeError(f"真实 AI 返回的关系解释不是合法 JSON：{exc}") from exc
    return parsed


def call_experience_summary_ai(manifest: Manifest, summary_context: dict[str, Any], raw_response_path: Path | None = None) -> dict[str, Any]:
    runtime = _resolve_ai_runtime(manifest)
    api_key = os.environ.get(runtime["api_key_env"])
    if not api_key:
        raise RuntimeError(f"缺少真实 AI Key：请在项目根目录 .env 里配置 {runtime['api_key_env']}=你的Key。")
    base_url = os.environ.get(runtime["base_url_env"], runtime["default_base_url"]).rstrip("/")
    model = os.environ.get(runtime["model_env"], runtime["default_model"])
    url = f"{base_url}/chat/completions"
    body = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "你是游戏数值策划的配表经验整理助手。只返回一个严格 JSON 对象，不要 Markdown。"
                    "目标是把用户随手写的经验整理成保存前可审核的知识。"
                    "JSON 必须包含：summary_title, review_text, activity_templates, field_mappings, "
                    "personal_rules, questions, risk_notes, conflicts。"
                    "review_text 用中文输出，适合用户直接编辑后保存；字段映射请尽量写成 `规划列名 -> table.field`。"
                    "如果上下文里有 existing_experiences，请逐条比较新经验和历史经验，只报告有证据的冲突。"
                    "conflicts 是数组，每项包含 conflict_type, severity, existing_experience_id, existing_title, "
                    "reason, new_value, existing_value, recommendation；没有冲突时返回空数组。"
                    "不要编造没有证据的配置值；不确定的内容放到 questions。"
                ),
            },
            {
                "role": "user",
                "content": json.dumps(summary_context, ensure_ascii=False),
            },
        ],
        "response_format": {"type": "json_object"},
        **runtime["extra_body"],
    }
    if runtime["temperature"] is not None:
        body["temperature"] = runtime["temperature"]
    request = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        message = exc.read().decode("utf-8", errors="replace")[:1000]
        raise RuntimeError(f"{runtime['label']} 经验整理请求失败：HTTP {exc.code} {message}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"{runtime['label']} 经验整理网络连接失败：{exc.reason}") from exc
    if raw_response_path:
        write_json(raw_response_path, {"provider": runtime["id"], "base_url": base_url, "model": model, "response": payload})
    content = payload["choices"][0]["message"]["content"]
    try:
        return json.loads(content)
    except json.JSONDecodeError as exc:
        if raw_response_path:
            write_text(raw_response_path.with_name("experience-summary-ai-invalid-content.txt"), content)
        raise RuntimeError(f"经验整理返回的内容不是合法 JSON：{exc}") from exc


def call_draft_diagnostics_ai(manifest: Manifest, diagnostic_context: dict[str, Any], raw_response_path: Path | None = None) -> dict[str, Any]:
    runtime = _resolve_ai_runtime(manifest)
    api_key = os.environ.get(runtime["api_key_env"])
    if not api_key:
        raise RuntimeError(f"缺少真实 AI Key：请在项目根目录 .env 里配置 {runtime['api_key_env']}=你的Key。")
    base_url = os.environ.get(runtime["base_url_env"], runtime["default_base_url"]).rstrip("/")
    model = os.environ.get(runtime["model_env"], runtime["default_model"])
    url = f"{base_url}/chat/completions"
    body = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You diagnose why a game meta configuration draft has zero operations. "
                    "Return only strict JSON with keys: summary, reasons, missing_information, "
                    "suggested_target_tables, suggested_field_mappings, next_steps. "
                    "Do not invent config values or patch operations. Use only the provided context."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(diagnostic_context, ensure_ascii=False),
            },
        ],
        "response_format": {"type": "json_object"},
        **runtime["extra_body"],
    }
    if runtime["temperature"] is not None:
        body["temperature"] = runtime["temperature"]
    request = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        message = exc.read().decode("utf-8", errors="replace")[:1000]
        raise RuntimeError(f"{runtime['label']} 空草案诊断请求失败：HTTP {exc.code} {message}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"{runtime['label']} 空草案诊断网络连接失败：{exc.reason}") from exc
    if raw_response_path:
        write_json(raw_response_path, {"provider": runtime["id"], "base_url": base_url, "model": model, "response": payload})
    content = payload["choices"][0]["message"]["content"]
    try:
        return json.loads(content)
    except json.JSONDecodeError as exc:
        if raw_response_path:
            write_text(raw_response_path.with_name("draft-diagnostics-ai-invalid-content.txt"), content)
        raise RuntimeError(f"空草案诊断返回的内容不是合法 JSON：{exc}") from exc


def call_case_review_ai(manifest: Manifest, review_context: dict[str, Any], raw_response_path: Path | None = None) -> dict[str, Any]:
    runtime = _resolve_ai_runtime(manifest)
    api_key = os.environ.get(runtime["api_key_env"])
    if not api_key:
        raise RuntimeError(f"缺少真实 AI Key：请在项目根目录 .env 里配置 {runtime['api_key_env']}=你的Key。")
    base_url = os.environ.get(runtime["base_url_env"], runtime["default_base_url"]).rstrip("/")
    model = os.environ.get(runtime["model_env"], runtime["default_model"])
    url = f"{base_url}/chat/completions"
    body = {
        "model": model,
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are a senior game configuration reviewer. The user has inspected Excel outputs and "
                    "listed problems with one AI-generated configuration run. Summarize the correction into "
                    "reusable lessons for future draft generation. Return only strict JSON with keys: "
                    "summary, mistakes, lessons, avoid_next_time, affected_tables, confidence. "
                    "Do not invent new config values. Use Chinese for user-facing text."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(review_context, ensure_ascii=False),
            },
        ],
        "response_format": {"type": "json_object"},
        **runtime["extra_body"],
    }
    if runtime["temperature"] is not None:
        body["temperature"] = runtime["temperature"]
    request = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        message = exc.read().decode("utf-8", errors="replace")[:1000]
        raise RuntimeError(f"{runtime['label']} 案例复盘请求失败：HTTP {exc.code} {message}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"{runtime['label']} 案例复盘网络连接失败：{exc.reason}") from exc
    if raw_response_path:
        write_json(raw_response_path, {"provider": runtime["id"], "base_url": base_url, "model": model, "response": payload})
    content = payload["choices"][0]["message"]["content"]
    try:
        return json.loads(content)
    except json.JSONDecodeError as exc:
        if raw_response_path:
            write_text(raw_response_path.with_name("case-review-ai-invalid-content.txt"), content)
        raise RuntimeError(f"案例复盘返回的内容不是合法 JSON：{exc}") from exc
