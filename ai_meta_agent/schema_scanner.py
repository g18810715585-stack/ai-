from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config_discovery import _iter_workbooks, _non_data_sheet_reason, _resolve
from .io_utils import write_json, write_text
from .models import Manifest


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def _detect_header(values: list[list[Any]]) -> tuple[int | None, list[str]]:
    best_index: int | None = None
    best_headers: list[str] = []
    best_score = 0
    for index, row in enumerate(values[:10]):
        headers = [str(_cell_value(value) or "") for value in row]
        non_empty = [item for item in headers if item]
        unique_count = len(set(non_empty))
        score = len(non_empty) + unique_count
        if len(non_empty) >= 2 and score > best_score:
            best_index = index
            best_headers = headers
            best_score = score
    return best_index, best_headers


def _infer_type(values: list[Any]) -> str:
    usable = [value for value in values if value not in (None, "")]
    if not usable:
        return "any"
    if all(isinstance(value, bool) for value in usable):
        return "bool"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in usable):
        return "int"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in usable):
        return "float"
    return "str"


def _infer_primary_key(headers: list[str], rows: list[dict[str, Any]], sheet_name: str) -> list[str]:
    candidates = []
    lower_sheet = sheet_name.lower()
    for header in headers:
        lower = header.lower()
        if lower == "id" or lower.endswith("_id") or lower.endswith("id"):
            candidates.append(header)
    preferred = [field for field in candidates if lower_sheet in field.lower() or field.lower() == f"{lower_sheet}_id"]
    for field in preferred + candidates:
        values = [row.get(field) for row in rows if row.get(field) not in (None, "")]
        if values and len(values) == len(set(values)):
            return [field]
    if headers:
        values = [row.get(headers[0]) for row in rows if row.get(headers[0]) not in (None, "")]
        if values and len(values) == len(set(values)):
            return [headers[0]]
    return []


def _sheet_profile(workbook_path: Path, sheet: Any, sample_limit: int) -> dict[str, Any] | None:
    rows_raw = [[_cell_value(cell.value) for cell in row] for row in sheet.iter_rows(max_row=min(sheet.max_row or 0, 200))]
    header_index, headers = _detect_header(rows_raw)
    if header_index is None:
        return None
    headers = [header for header in headers if header]
    if len(headers) < 2:
        return None
    data_rows: list[dict[str, Any]] = []
    for raw in rows_raw[header_index + 1 :]:
        row = {header: raw[idx] if idx < len(raw) else None for idx, header in enumerate(headers)}
        if any(value not in (None, "") for value in row.values()):
            data_rows.append(row)
        if len(data_rows) >= sample_limit:
            break
    field_values: dict[str, list[Any]] = defaultdict(list)
    for row in data_rows:
        for field, value in row.items():
            field_values[field].append(value)
    fields = {
        field: {
            "type": _infer_type(field_values[field]),
            "required": bool(field_values[field]) and all(value not in (None, "") for value in field_values[field]),
        }
        for field in headers
    }
    return {
        "table": sheet.title,
        "sheet": sheet.title,
        "source_file": str(workbook_path),
        "header_row": header_index + 1,
        "sample_count": len(data_rows),
        "primary_key": _infer_primary_key(headers, data_rows, sheet.title),
        "fields": fields,
        "sample_rows": data_rows,
    }


def scan_config_schema(manifest: Manifest, base_dir: Path, run_dir: Path, sample_limit: int = 5) -> dict[str, Any]:
    roots = []
    tables: dict[str, Any] = {}
    skipped_sheets: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    duplicate_sheets: dict[str, list[str]] = defaultdict(list)
    for root in manifest.config_roots:
        root_path = _resolve(base_dir, root.path)
        workbooks = _iter_workbooks(root_path, root.recursive)
        roots.append({"path": str(root_path), "recursive": root.recursive, "workbooks": len(workbooks)})
        for workbook_path in workbooks:
            try:
                workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            except Exception as exc:  # noqa: BLE001 - report bad workbook and continue scanning.
                errors.append({"path": str(workbook_path), "message": str(exc)})
                continue
            try:
                for sheet in workbook.worksheets:
                    reason = _non_data_sheet_reason(sheet)
                    if reason:
                        skipped_sheets.append(
                            {
                                "path": str(workbook_path),
                                "name": sheet.title,
                                "max_row": sheet.max_row,
                                "max_column": sheet.max_column,
                                "reason": reason,
                            }
                        )
                        continue
                    profile = _sheet_profile(workbook_path, sheet, sample_limit)
                    if not profile:
                        skipped_sheets.append(
                            {
                                "path": str(workbook_path),
                                "name": sheet.title,
                                "max_row": sheet.max_row,
                                "max_column": sheet.max_column,
                                "reason": "no_header",
                            }
                        )
                        continue
                    table_name = profile["table"]
                    duplicate_sheets[table_name].append(str(workbook_path))
                    if table_name not in tables:
                        tables[table_name] = {
                            "sheet": table_name,
                            "source_file": profile["source_file"],
                            "primary_key": profile["primary_key"],
                            "overwrite_strategy": "field_level",
                            "ai_write_permission": "supervised_write",
                            "allow_update_fields": list(profile["fields"].keys()),
                            "preserve_fields": [],
                            "block_update_fields": profile["primary_key"],
                            "fields": profile["fields"],
                            "field_aliases": {},
                            "sample_rows": profile["sample_rows"],
                        }
            finally:
                workbook.close()
    duplicates = {name: paths for name, paths in duplicate_sheets.items() if len(paths) > 1}
    reference_candidates = _reference_candidates(tables)
    schema_draft = {
        "version": 1,
        "tables": {
            name: {key: value for key, value in table.items() if key not in {"sample_rows", "source_file"}}
            for name, table in sorted(tables.items())
        },
        "risk": {
            "auto_apply_confidence": 0.9,
            "needs_review_confidence": 0.7,
            "large_delete_threshold": 20,
            "large_update_threshold": 50,
        },
    }
    report = {
        "roots": roots,
        "table_count": len(tables),
        "tables": tables,
        "duplicates": duplicates,
        "reference_candidates": reference_candidates,
        "skipped_sheets": skipped_sheets,
        "errors": errors,
    }
    write_json(run_dir / "schema-draft.json", schema_draft)
    write_json(run_dir / "schema-scan.json", report)
    write_text(run_dir / "schema-scan.md", _schema_scan_markdown(report))
    return {"schema_draft": schema_draft, "report": report}


def _reference_candidates(tables: dict[str, Any]) -> list[dict[str, str]]:
    primary_by_field = {}
    for table_name, table in tables.items():
        for field in table.get("primary_key", []):
            primary_by_field[field] = table_name
    candidates = []
    for table_name, table in tables.items():
        for field in table.get("fields", {}):
            if field in primary_by_field and primary_by_field[field] != table_name:
                candidates.append({"from_table": table_name, "field": field, "to_table": primary_by_field[field]})
    return candidates


def _schema_scan_markdown(report: dict[str, Any]) -> str:
    lines = ["# Schema Scan", ""]
    lines.append("## Roots")
    for root in report["roots"]:
        lines.append(f"- {root['path']} recursive={root['recursive']} workbooks={root['workbooks']}")
    lines.append("")
    lines.append(f"## Tables ({report['table_count']})")
    for table_name, table in sorted(report["tables"].items()):
        lines.append(f"- `{table_name}` fields={len(table['fields'])} pk={table['primary_key']} source={table['source_file']}")
    lines.append("")
    if report["duplicates"]:
        lines.append("## Duplicate Sheet Names")
        for name, paths in sorted(report["duplicates"].items()):
            lines.append(f"- `{name}` appears in {len(paths)} workbooks")
        lines.append("")
    lines.append(f"## Skipped Sheets ({len(report['skipped_sheets'])})")
    counts = Counter(item["reason"] for item in report["skipped_sheets"])
    for reason, count in sorted(counts.items()):
        lines.append(f"- {reason}: {count}")
    lines.append("")
    if report["errors"]:
        lines.append("## Errors")
        for error in report["errors"][:20]:
            lines.append(f"- {error['path']}: {error['message']}")
        lines.append("")
    return "\n".join(lines)
