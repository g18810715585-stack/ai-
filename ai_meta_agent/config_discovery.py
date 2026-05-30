from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ConfigTableRef, Manifest, SchemaBundle


EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
NON_DATA_SHEET_EXACT = {
    "sheet",
    "sheet1",
    "readme",
    "index",
    "contents",
    "changelog",
    "说明",
    "配置说明",
    "数据说明",
    "目录",
    "索引",
    "封面",
    "首页",
    "备注",
    "示例",
    "样例",
    "模板",
    "更新记录",
    "修改记录",
    "版本记录",
    "废弃",
}
NON_DATA_SHEET_KEYWORDS = (
    "说明",
    "目录",
    "备注",
    "示例",
    "样例",
    "模板",
    "更新记录",
    "修改记录",
    "版本记录",
    "readme",
    "changelog",
    "deprecated",
)


def _resolve(base_dir: Path, value: str) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return (base_dir / path).resolve()


def _iter_workbooks(root: Path, recursive: bool) -> list[Path]:
    if root.is_file() and root.suffix.lower() in EXCEL_SUFFIXES:
        return [root]
    if not root.exists() or not root.is_dir():
        return []
    pattern = "**/*" if recursive else "*"
    return sorted(path for path in root.glob(pattern) if path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES and not path.name.startswith("~$"))


def _non_data_sheet_reason(sheet: Any) -> str | None:
    normalized = sheet.title.strip().lower()
    if getattr(sheet, "sheet_state", "visible") != "visible":
        return "hidden_sheet"
    if normalized in NON_DATA_SHEET_EXACT:
        return "non_data_name"
    if any(keyword in normalized for keyword in NON_DATA_SHEET_KEYWORDS):
        return "non_data_keyword"
    if (sheet.max_row or 0) < 2 or (sheet.max_column or 0) < 2:
        return "too_small"
    return None


def _workbook_sheet_info(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    data_sheets: list[dict[str, Any]] = []
    skipped_sheets: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            info = {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            }
            reason = _non_data_sheet_reason(sheet)
            if reason:
                skipped_sheets.append({**info, "reason": reason})
            else:
                data_sheets.append(info)
    finally:
        workbook.close()
    return data_sheets, skipped_sheets


def discover_config_tables(manifest: Manifest, schema: SchemaBundle, base_dir: Path) -> tuple[Manifest, dict[str, Any]]:
    resolved = dict(manifest.config_tables)
    diagnostics: dict[str, Any] = {
        "roots": [],
        "matched": {},
        "unmatched_tables": [],
        "skipped_sheets": [],
        "errors": [],
    }
    if not manifest.config_roots:
        diagnostics["unmatched_tables"] = [table for table in schema.tables if table not in resolved]
        return manifest, diagnostics

    workbook_cache: dict[Path, list[dict[str, Any]]] = {}
    candidates: list[Path] = []
    for root in manifest.config_roots:
        root_path = _resolve(base_dir, root.path)
        workbooks = _iter_workbooks(root_path, root.recursive)
        diagnostics["roots"].append({"path": str(root_path), "recursive": root.recursive, "workbooks": len(workbooks)})
        candidates.extend(workbooks)

    for table_name, table in schema.tables.items():
        if table_name in resolved:
            diagnostics["matched"][table_name] = {
                "path": resolved[table_name].path,
                "sheet": resolved[table_name].sheet,
                "source": "manifest.config_tables",
            }
            continue
        expected_sheet = table.sheet or table_name
        expected_names = {table_name.lower(), expected_sheet.lower()}
        for workbook_path in candidates:
            try:
                if workbook_path not in workbook_cache:
                    data_sheets, skipped_sheets = _workbook_sheet_info(workbook_path)
                    workbook_cache[workbook_path] = data_sheets
                    for skipped in skipped_sheets:
                        diagnostics["skipped_sheets"].append({"path": str(workbook_path), **skipped})
                sheet_infos = workbook_cache[workbook_path]
            except Exception as exc:  # noqa: BLE001 - report bad workbook, keep scanning others.
                diagnostics["errors"].append({"path": str(workbook_path), "message": str(exc)})
                continue
            matched_sheet = next((sheet["name"] for sheet in sheet_infos if sheet["name"].lower() in expected_names), None)
            if matched_sheet:
                resolved[table_name] = ConfigTableRef(path=str(workbook_path), sheet=matched_sheet)
                diagnostics["matched"][table_name] = {"path": str(workbook_path), "sheet": matched_sheet, "source": "sheet_name"}
                break
        if table_name not in resolved:
            diagnostics["unmatched_tables"].append(table_name)

    manifest.config_tables = resolved
    return manifest, diagnostics
