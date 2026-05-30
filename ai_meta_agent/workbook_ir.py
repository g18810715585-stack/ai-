from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .feishu import read_feishu_source
from .models import CellIR, PlanningSource, SheetIR, SourceKind, WorkbookIR


def _cell_fill(cell: Any) -> str | None:
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None
    rgb = fill.fgColor.rgb
    if not rgb or rgb == "00000000":
        return None
    return str(rgb)


def _merged_lookup(sheet: Any) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for merged in sheet.merged_cells.ranges:
        label = str(merged)
        for row in sheet.iter_rows(min_row=merged.min_row, max_row=merged.max_row, min_col=merged.min_col, max_col=merged.max_col):
            for cell in row:
                lookup[cell.coordinate] = label
    return lookup


def _detect_header(sheet: Any) -> tuple[int | None, list[str]]:
    best_row: int | None = None
    best_values: list[str] = []
    best_count = 0
    scan_rows = min(sheet.max_row, 20)
    for row_idx in range(1, scan_rows + 1):
        values = []
        count = 0
        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(row_idx, col_idx).value
            label = "" if value is None else str(value).strip()
            values.append(label)
            if label:
                count += 1
        if count > best_count:
            best_row = row_idx
            best_values = values
            best_count = count
    if best_count < 2:
        return None, []
    return best_row, best_values


def _sample_rows(sheet: Any, header_row: int | None, headers: list[str], limit: int = 20) -> list[dict[str, Any]]:
    if not header_row or not headers:
        return []
    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, min(sheet.max_row, header_row + limit) + 1):
        item: dict[str, Any] = {"__row": row_idx}
        non_empty = False
        for col_idx, header in enumerate(headers, start=1):
            if not header:
                continue
            value = sheet.cell(row_idx, col_idx).value
            item[header] = value
            if value not in (None, ""):
                non_empty = True
        if non_empty:
            rows.append(item)
    return rows


def _detect_header_from_matrix(values: list[list[Any]]) -> tuple[int | None, list[str]]:
    best_row: int | None = None
    best_values: list[str] = []
    best_count = 0
    for row_idx, row in enumerate(values[:20], start=1):
        labels = ["" if value is None else str(value).strip() for value in row]
        count = sum(1 for label in labels if label)
        if count > best_count:
            best_row = row_idx
            best_values = labels
            best_count = count
    if best_count < 2:
        return None, []
    return best_row, best_values


def _sample_rows_from_matrix(values: list[list[Any]], header_row: int | None, headers: list[str], limit: int = 20) -> list[dict[str, Any]]:
    if not header_row or not headers:
        return []
    rows: list[dict[str, Any]] = []
    for row_idx, raw in enumerate(values[header_row : header_row + limit], start=header_row + 1):
        item: dict[str, Any] = {"__row": row_idx}
        non_empty = False
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            value = raw[col_idx] if col_idx < len(raw) else None
            item[header] = value
            if value not in (None, ""):
                non_empty = True
        if non_empty:
            rows.append(item)
    return rows


def _matrix_to_sheet_ir(name: str, values: list[list[Any]], max_cells_per_sheet: int = 400) -> SheetIR:
    max_row = len(values)
    max_column = max((len(row) for row in values), default=0)
    header_row, headers = _detect_header_from_matrix(values)
    cells: list[CellIR] = []
    added = 0
    for row_idx, row in enumerate(values, start=1):
        for col_idx, value in enumerate(row, start=1):
            if added >= max_cells_per_sheet:
                break
            if value in (None, ""):
                continue
            cells.append(
                CellIR(
                    row=row_idx,
                    column=col_idx,
                    address=f"{get_column_letter(col_idx)}{row_idx}",
                    value=value,
                )
            )
            added += 1
        if added >= max_cells_per_sheet:
            break
    return SheetIR(
        name=name,
        max_row=max_row,
        max_column=max_column,
        headers=headers,
        header_row=header_row,
        sample_rows=_sample_rows_from_matrix(values, header_row, headers),
        cells=cells,
    )


def _text_to_sheet_ir(name: str, text: str, max_lines: int = 300) -> SheetIR:
    rows: list[list[Any]] = [["section", "content"]]
    for index, line in enumerate((line.strip() for line in text.splitlines()), start=1):
        if not line:
            continue
        rows.append([index, line])
        if len(rows) >= max_lines:
            break
    if len(rows) == 1 and text.strip():
        rows.append([1, text.strip()])
    return _matrix_to_sheet_ir(name, rows)


def load_local_workbook_ir(source: PlanningSource, path: Path, max_cells_per_sheet: int = 400) -> WorkbookIR:
    workbook = load_workbook(path, data_only=True)
    sheets: list[SheetIR] = []
    for sheet in workbook.worksheets:
        merged_lookup = _merged_lookup(sheet)
        hidden_rows = [idx for idx, dim in sheet.row_dimensions.items() if dim.hidden]
        hidden_columns = [col for col, dim in sheet.column_dimensions.items() if dim.hidden]
        header_row, headers = _detect_header(sheet)
        cells: list[CellIR] = []
        added = 0
        for row in sheet.iter_rows():
            for cell in row:
                if added >= max_cells_per_sheet:
                    break
                if cell.value is None and cell.coordinate not in merged_lookup:
                    continue
                column_letter = get_column_letter(cell.column)
                cells.append(
                    CellIR(
                        row=cell.row,
                        column=cell.column,
                        address=cell.coordinate,
                        value=cell.value,
                        is_merged=cell.coordinate in merged_lookup,
                        merged_range=merged_lookup.get(cell.coordinate),
                        is_hidden_row=cell.row in hidden_rows,
                        is_hidden_column=column_letter in hidden_columns,
                        fill=_cell_fill(cell),
                        comment=cell.comment.text if cell.comment else None,
                    )
                )
                added += 1
            if added >= max_cells_per_sheet:
                break
        sheets.append(
            SheetIR(
                name=sheet.title,
                max_row=sheet.max_row,
                max_column=sheet.max_column,
                hidden_rows=hidden_rows,
                hidden_columns=hidden_columns,
                merged_ranges=[str(item) for item in sheet.merged_cells.ranges],
                headers=headers,
                header_row=header_row,
                sample_rows=_sample_rows(sheet, header_row, headers),
                cells=cells,
            )
        )
    return WorkbookIR(source_id=source.id, source_type=source.kind, path=str(path), sheets=sheets)


def load_feishu_workbook_ir(source: PlanningSource, base_dir: Path) -> WorkbookIR:
    payload = read_feishu_source(source.url or "", base_dir, sheet_id=source.sheet_id, range_name=source.range)
    if payload.kind == "sheet":
        sheet_name = payload.title or payload.sheet_id or source.role or "飞书规划表"
        sheets = [_matrix_to_sheet_ir(sheet_name, payload.values)]
    else:
        sheets = [_text_to_sheet_ir(payload.title or "飞书规划文档", payload.text)]
    return WorkbookIR(source_id=source.id, source_type=source.kind, url=source.url, sheets=sheets)


def load_source_ir(source: PlanningSource, base_dir: Path) -> WorkbookIR:
    if source.kind == SourceKind.LOCAL_EXCEL:
        path = Path(source.path or "")
        if not path.is_absolute():
            path = (base_dir / path).resolve()
        if not path.exists():
            raise FileNotFoundError(f"Planning workbook not found: {path}")
        return load_local_workbook_ir(source, path)
    return load_feishu_workbook_ir(source, base_dir)
