from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .models import Manifest, Patch, PatchOperation, SchemaBundle, ValidationIssue, ValidationReport


@dataclass
class WorkbookState:
    source_path: Path
    preview_path: Path
    workbook: Workbook


def _headers(sheet: Any) -> list[str]:
    return list(_header_index(sheet).keys())


def _header_index(sheet: Any) -> dict[str, int]:
    index: dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(1, col).value
        if value not in (None, ""):
            index[str(value)] = col
    return index


def _ensure_sheet(workbook: Workbook, sheet_name: str, fields: list[str]) -> Any:
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    index = _header_index(sheet)
    if not index:
        for col, field in enumerate(fields, start=1):
            sheet.cell(1, col).value = field
            index[field] = col
    else:
        next_col = max(index.values(), default=0) + 1
        for field in fields:
            if field not in index:
                sheet.cell(1, next_col).value = field
                index[field] = next_col
                next_col += 1
    return sheet


def _row_dict(sheet: Any, row_idx: int, index: dict[str, int] | None = None, fields: list[str] | None = None) -> dict[str, Any]:
    current_index = index or _header_index(sheet)
    selected_fields = fields or list(current_index.keys())
    return {
        field: sheet.cell(row_idx, current_index[field]).value if field in current_index else None
        for field in selected_fields
    }


def _row_matches(row: dict[str, Any], match: dict[str, Any]) -> bool:
    return all(str(row.get(field)) == str(value) for field, value in match.items())


def _find_rows(sheet: Any, match: dict[str, Any], index: dict[str, int] | None = None) -> list[int]:
    current_index = index or _header_index(sheet)
    if any(field not in current_index for field in match):
        return []
    rows: list[int] = []
    match_fields = list(match.keys())
    for row_idx in range(2, sheet.max_row + 1):
        row = _row_dict(sheet, row_idx, current_index, match_fields)
        if _row_matches(row, match):
            rows.append(row_idx)
    return rows


def _write_row(sheet: Any, row_idx: int, row: dict[str, Any], index: dict[str, int] | None = None) -> dict[str, int]:
    current_index = index if index is not None else _header_index(sheet)
    next_col = max(current_index.values(), default=0) + 1
    for field, value in row.items():
        if field not in current_index:
            sheet.cell(1, next_col).value = field
            current_index[field] = next_col
            next_col += 1
        sheet.cell(row_idx, current_index[field]).value = value
    return current_index


def _table_ref(manifest: Manifest, schema: SchemaBundle, table_name: str, base_dir: Path) -> tuple[Path, str]:
    ref = manifest.config_tables.get(table_name)
    table = schema.tables[table_name]
    if not ref:
        raise KeyError(f"Manifest missing config table path for {table_name}")
    source = Path(ref.path)
    if not source.is_absolute():
        source = (base_dir / source).resolve()
    return source, ref.sheet or table.sheet or table_name


def _open_states(manifest: Manifest, schema: SchemaBundle, patch: Patch, base_dir: Path, run_dir: Path) -> dict[str, WorkbookState]:
    states: dict[str, WorkbookState] = {}
    preview_dir = run_dir / "preview"
    preview_dir.mkdir(parents=True, exist_ok=True)
    for operation in patch.operations:
        source_path, _ = _table_ref(manifest, schema, operation.target_table, base_dir)
        key = str(source_path)
        if key in states:
            continue
        if not source_path.exists():
            workbook = Workbook()
            workbook.active.title = "Sheet1"
        else:
            workbook = load_workbook(source_path)
        preview_path = preview_dir / f"{source_path.stem}.preview{source_path.suffix or '.xlsx'}"
        states[key] = WorkbookState(source_path=source_path, preview_path=preview_path, workbook=workbook)
    return states


def _assert_overwrite_targets_writable(states: dict[str, WorkbookState]) -> None:
    locked: list[str] = []
    for state in states.values():
        if not state.source_path.exists():
            continue
        try:
            with state.source_path.open("r+b"):
                pass
        except OSError as exc:
            locked.append(f"{state.source_path} ({exc})")
    if locked:
        joined = "\n".join(f"- {item}" for item in locked)
        raise PermissionError(f"原表暂时无法写入，请先关闭 Excel 或解除文件占用后重试：\n{joined}")


def _normalize_key(row: dict[str, Any], primary_key: list[str]) -> tuple[Any, ...]:
    return tuple(row.get(field) for field in primary_key)


def _row_key(row: dict[str, Any], primary_key: list[str], fallback: int) -> list[Any]:
    if not primary_key:
        return [fallback]
    return list(_normalize_key(row, primary_key))


def _table_diff(diffs: dict[str, Any], file_key: str, sheet_name: str) -> dict[str, list[Any]]:
    file_diff = diffs.setdefault(file_key, {})
    return file_diff.setdefault(sheet_name, {"inserted": [], "deleted": [], "changed": []})


def _track_table(tables_by_file: dict[str, set[str]], file_key: str, table_name: str) -> None:
    tables_by_file.setdefault(file_key, set()).add(table_name)


def _track_row(rows_by_file: dict[str, dict[str, set[int]]], file_key: str, table_name: str, row_idx: int | None = None) -> None:
    table_rows = rows_by_file.setdefault(file_key, {}).setdefault(table_name, set())
    if row_idx is not None:
        table_rows.add(row_idx)


def _validate_workbook(
    workbook: Workbook,
    schema: SchemaBundle,
    table_names: set[str] | None = None,
    touched_rows: dict[str, set[int]] | None = None,
) -> ValidationReport:
    report = ValidationReport()
    selected_tables = table_names if table_names is not None else set(schema.tables.keys())
    for table_name in selected_tables:
        table = schema.tables.get(table_name)
        if not table:
            continue
        sheet_name = table.sheet or table_name
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        index = _header_index(sheet)
        headers = set(index.keys())
        for field, spec in table.fields.items():
            if spec.required and field not in headers:
                report.errors.append(ValidationIssue(level="error", table=table_name, field=field, message="required field missing from sheet"))

        if touched_rows is not None and table_name in touched_rows:
            rows_to_check = sorted(row_idx for row_idx in touched_rows[table_name] if row_idx <= sheet.max_row)
        else:
            rows_to_check = list(range(2, sheet.max_row + 1))

        field_names = list(table.fields.keys())
        for row_idx in rows_to_check:
            row = _row_dict(sheet, row_idx, index, field_names)
            if not any(value not in (None, "") for value in row.values()):
                continue
            for field, spec in table.fields.items():
                value = row.get(field)
                if spec.required and value in (None, ""):
                    report.errors.append(ValidationIssue(level="error", table=table_name, row=row_idx, field=field, message="required value missing"))
                if value not in (None, "") and spec.type == "int":
                    try:
                        int(value)
                    except (TypeError, ValueError):
                        report.errors.append(ValidationIssue(level="error", table=table_name, row=row_idx, field=field, message="expected int"))

        touched_for_table = touched_rows.get(table_name, set()) if touched_rows is not None else None
        if table.primary_key and all(field in index for field in table.primary_key) and (touched_for_table is None or touched_for_table):
            seen: dict[tuple[Any, ...], int] = {}
            touched_keys: set[tuple[Any, ...]] = set()
            for row_idx in range(2, sheet.max_row + 1):
                key = tuple(sheet.cell(row_idx, index[field]).value for field in table.primary_key)
                if not any(value not in (None, "") for value in key):
                    continue
                if touched_for_table is not None and row_idx in touched_for_table:
                    touched_keys.add(key)
                if key in seen:
                    if touched_for_table is None or row_idx in touched_for_table or key in touched_keys:
                        report.errors.append(ValidationIssue(level="error", table=table_name, row=row_idx, message=f"duplicate primary key {key}"))
                else:
                    seen[key] = row_idx
    return report


def apply_patch(manifest: Manifest, schema: SchemaBundle, patch: Patch, base_dir: Path, run_dir: Path, write_mode: str = "preview") -> dict[str, Any]:
    if write_mode not in {"preview", "overwrite"}:
        raise ValueError("write_mode must be preview or overwrite")
    states = _open_states(manifest, schema, patch, base_dir, run_dir)
    if write_mode == "overwrite":
        _assert_overwrite_targets_writable(states)
    rollback_ops: list[PatchOperation] = []
    operation_results: list[dict[str, Any]] = []
    touched_tables_by_file: dict[str, set[str]] = {}
    touched_rows_by_file: dict[str, dict[str, set[int]]] = {}
    diffs: dict[str, Any] = {}

    for operation in patch.operations:
        table = schema.tables[operation.target_table]
        source_path, sheet_name = _table_ref(manifest, schema, operation.target_table, base_dir)
        file_key = str(source_path)
        state = states[file_key]
        sheet = _ensure_sheet(state.workbook, sheet_name, list(table.fields.keys()))
        index = _header_index(sheet)
        table_diff = _table_diff(diffs, file_key, sheet_name)
        _track_table(touched_tables_by_file, file_key, operation.target_table)
        _track_row(touched_rows_by_file, file_key, operation.target_table)

        if operation.op == "update":
            rows = _find_rows(sheet, operation.match, index)
            old_rows = [_row_dict(sheet, row_idx, index) for row_idx in rows]
            allowed = {
                field: value
                for field, value in operation.set.items()
                if field not in table.block_update_fields and (not table.allow_update_fields or field in table.allow_update_fields)
            }
            for row_idx, before in zip(rows, old_rows):
                _write_row(sheet, row_idx, allowed, index)
                after = _row_dict(sheet, row_idx, index)
                if before != after:
                    table_diff["changed"].append(
                        {
                            "row": row_idx,
                            "key": _row_key(after, table.primary_key, row_idx),
                            "before": before,
                            "after": after,
                        }
                    )
                _track_row(touched_rows_by_file, file_key, operation.target_table, row_idx)
            for old in old_rows:
                rollback_ops.append(
                    PatchOperation(
                        op="update",
                        target_table=operation.target_table,
                        match=operation.match,
                        set=old,
                        source_ref=operation.source_ref,
                        reason=f"rollback for {operation.reason}",
                        confidence=1.0,
                        risk_level="low",
                        needs_confirmation=False,
                    )
                )
            operation_results.append({"op": operation.op, "target_table": operation.target_table, "affected_rows": len(rows)})

        elif operation.op == "insert":
            for row in operation.rows:
                row_idx = sheet.max_row + 1
                _write_row(sheet, row_idx, row, index)
                inserted = _row_dict(sheet, row_idx, index)
                table_diff["inserted"].append(inserted)
                _track_row(touched_rows_by_file, file_key, operation.target_table, row_idx)
                match = {field: row.get(field) for field in table.primary_key if field in row}
                if match:
                    rollback_ops.append(
                        PatchOperation(
                            op="delete_where",
                            target_table=operation.target_table,
                            match=match,
                            source_ref=operation.source_ref,
                            reason=f"rollback insert for {operation.reason}",
                            confidence=1.0,
                            risk_level="low",
                            needs_confirmation=False,
                        )
                    )
            operation_results.append({"op": operation.op, "target_table": operation.target_table, "affected_rows": len(operation.rows)})

        elif operation.op == "delete_where":
            rows = _find_rows(sheet, operation.match, index)
            old_rows = [_row_dict(sheet, row_idx, index) for row_idx in rows]
            table_diff["deleted"].extend(old_rows)
            for row_idx in sorted(rows, reverse=True):
                sheet.delete_rows(row_idx, 1)
            if old_rows:
                rollback_ops.append(
                    PatchOperation(
                        op="insert",
                        target_table=operation.target_table,
                        rows=old_rows,
                        source_ref=operation.source_ref,
                        reason=f"rollback delete for {operation.reason}",
                        confidence=1.0,
                        risk_level="low",
                        needs_confirmation=False,
                    )
                )
            operation_results.append({"op": operation.op, "target_table": operation.target_table, "affected_rows": len(rows)})

        elif operation.op == "replace_group":
            rows = _find_rows(sheet, operation.match, index)
            old_rows = [_row_dict(sheet, row_idx, index) for row_idx in rows]
            table_diff["deleted"].extend(old_rows)
            for row_idx in sorted(rows, reverse=True):
                sheet.delete_rows(row_idx, 1)
            for row in operation.rows:
                row_idx = sheet.max_row + 1
                _write_row(sheet, row_idx, row, index)
                inserted = _row_dict(sheet, row_idx, index)
                table_diff["inserted"].append(inserted)
                _track_row(touched_rows_by_file, file_key, operation.target_table, row_idx)
            rollback_ops.append(
                PatchOperation(
                    op="replace_group",
                    target_table=operation.target_table,
                    match=operation.match,
                    rows=old_rows,
                    source_ref=operation.source_ref,
                    reason=f"rollback replace group for {operation.reason}",
                    confidence=1.0,
                    risk_level="low",
                    needs_confirmation=False,
                )
            )
            operation_results.append({"op": operation.op, "target_table": operation.target_table, "affected_rows": len(rows) + len(operation.rows)})

    previews: dict[str, str] = {}
    backups: dict[str, str] = {}
    written_files: dict[str, str] = {}
    validation_reports: dict[str, Any] = {}
    for key, state in states.items():
        if state.source_path.exists():
            backup = run_dir / "backups" / state.source_path.name
            backup.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(state.source_path, backup)
            backups[key] = str(backup)
        state.workbook.save(state.preview_path)
        previews[key] = str(state.preview_path)
        if write_mode == "overwrite":
            state.source_path.parent.mkdir(parents=True, exist_ok=True)
            state.workbook.save(state.source_path)
            written_files[key] = str(state.source_path)
        report = _validate_workbook(
            state.workbook,
            schema,
            touched_tables_by_file.get(key, set()),
            touched_rows_by_file.get(key, {}),
        )
        validation_reports[key] = report.model_dump()
        diffs.setdefault(key, {})

    rollback = Patch(
        patch_id=f"rollback_{patch.patch_id}",
        project=patch.project,
        mode=patch.mode,
        operations=list(reversed(rollback_ops)),
        generated_by="ai-meta-agent",
    )
    return {
        "patch_id": patch.patch_id,
        "write_mode": write_mode,
        "operation_results": operation_results,
        "previews": previews,
        "backups": backups,
        "written_files": written_files,
        "diff": diffs,
        "validation": validation_reports,
        "rollback_patch": rollback.model_dump(mode="json", exclude_none=True),
    }
