from __future__ import annotations

from collections import Counter, deque
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import Manifest, SchemaBundle, TableSchema, resolve_path

PROFILE_SCAN_LIMIT = 5000
TAIL_ROW_LIMIT = 12
SAMPLE_VALUE_LIMIT = 8
ENUM_VALUE_LIMIT = 12
PROFILE_FIELD_LIMIT = 80

REFERENCE_KEYWORDS = (
    "id",
    "key",
    "list",
    "form",
    "group",
    "reward",
    "goods",
    "item",
    "task",
    "activity",
    "shop",
)
REFERENCE_WORDS = (
    "ID",
    "编号",
    "序号",
    "组",
    "奖励",
    "商品",
    "道具",
    "物品",
    "文案",
    "标题",
    "描述",
    "列表",
    "活动",
    "任务",
    "商店",
    "礼包",
    "积分",
    "条件",
    "类型",
    "状态",
    "开关",
)
GROUP_WORDS = ("组", "group", "group_id")
ENUM_HINTS = ("type", "status", "state", "is_", "类型", "状态", "开关", "是否")
GENERATED_STRATEGIES = {"new", "new_or_reuse"}


def build_target_table_profiles(
    manifest: Manifest,
    schema: SchemaBundle,
    base_dir: Path,
    field_dictionary: list[dict[str, Any]] | None = None,
    max_rows: int = PROFILE_SCAN_LIMIT,
) -> dict[str, Any]:
    """Build compact source-table context for AI and deterministic patch helpers.

    The profile is intentionally a summary, not a copy of the real config table:
    it includes ID/group baselines, low-cardinality enums, field samples, and
    tail rows so the draft step can understand current table state without
    sending whole workbooks to the model.
    """
    dictionary_index = _dictionary_index(field_dictionary or [])
    profiles: dict[str, Any] = {}
    for table_name, table in schema.tables.items():
        ref = manifest.config_tables.get(table_name)
        if not ref:
            continue
        path = resolve_path(base_dir, ref.path)
        if not path.exists():
            continue
        try:
            profiles[table_name] = _profile_table(
                path,
                ref.sheet or table.sheet or table_name,
                table_name,
                table,
                dictionary_index,
                max_rows,
            )
        except Exception as exc:  # noqa: BLE001 - profiles are advisory context only.
            profiles[table_name] = {"path": str(path), "sheet": ref.sheet or table.sheet or table_name, "error": str(exc)}
    return profiles


def _profile_table(
    path: Path,
    sheet_name: str,
    table_name: str,
    table: TableSchema,
    dictionary_index: dict[tuple[str, str], dict[str, Any]],
    max_rows: int,
) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        schema_fields = list(table.fields.keys())
        header_row, headers = _detect_header(sheet, schema_fields)
        if not header_row:
            return {"path": str(path), "sheet": sheet.title, "error": "未识别到表头"}

        header_to_index = {header: index for index, header in enumerate(headers) if header}
        profile_fields = _profile_fields(table_name, table, schema_fields, dictionary_index)
        field_stats = {
            field: _empty_field_stat(field, table_name, table, dictionary_index)
            for field in profile_fields
            if field in header_to_index
        }
        tail_rows: deque[dict[str, Any]] = deque(maxlen=TAIL_ROW_LIMIT)
        scanned_rows = 0
        data_rows_seen = 0
        scan_limit_reached = False
        for row_index, raw in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(value not in (None, "") for value in raw):
                continue
            if scanned_rows >= max_rows:
                scan_limit_reached = True
                break
            data_rows_seen += 1
            scanned_rows += 1
            compact_row: dict[str, Any] = {"__row": row_index}
            for field, stat in field_stats.items():
                index = header_to_index[field]
                value = _cell_value(raw[index] if index < len(raw) else None)
                if value in (None, ""):
                    continue
                compact_row[field] = value
                _update_field_stat(stat, value, row_index)
            if len(compact_row) > 1:
                tail_rows.append(compact_row)

        fields = {field: _finalize_field_stat(stat, scanned_rows) for field, stat in field_stats.items()}
        allocatable_fields = [
            field
            for field, stat in fields.items()
            if stat.get("allocation_role") and stat.get("next_value") is not None
        ]
        next_values = {field: fields[field]["next_value"] for field in allocatable_fields}
        enum_fields = [field for field, stat in fields.items() if stat.get("enum_values")]
        lookup_fields = [
            field
            for field, stat in fields.items()
            if "lookup_ref" in stat.get("roles", []) or ("foreign_key_candidate" in stat.get("roles", []) and not stat.get("allocation_role"))
        ]
        row_count_estimate = max((sheet.max_row or header_row) - header_row, data_rows_seen)

        return {
            "path": str(path),
            "sheet": sheet.title,
            "header_row": header_row,
            "row_count": row_count_estimate,
            "scanned_row_count": scanned_rows,
            "scan_limit": max_rows,
            "scan_limit_reached": scan_limit_reached or row_count_estimate > scanned_rows,
            "primary_key": table.primary_key,
            "group_key": table.group_key,
            "profile_fields": list(fields.keys()),
            "generation_summary": {
                "allocatable_fields": allocatable_fields,
                "lookup_fields": lookup_fields,
                "enum_fields": enum_fields,
                "note": "只对主键、组字段或字段字典标记为 new/new_or_reuse 的字段做确定性递增；外键类字段只作为查找上下文。",
            },
            "fields": fields,
            "next_values": next_values,
            "tail_rows": list(tail_rows),
        }
    finally:
        workbook.close()


def _detect_header(sheet: Any, schema_fields: list[str]) -> tuple[int | None, list[str]]:
    schema_set = set(schema_fields)
    best_row: int | None = None
    best_values: list[str] = []
    best_score = 0
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 0, 30), values_only=True), start=1):
        values = ["" if value is None else str(value).strip() for value in row]
        non_empty = sum(1 for value in values if value)
        matches = sum(1 for value in values if value in schema_set)
        score = matches * 10 + non_empty
        if score > best_score:
            best_row = row_index
            best_values = values
            best_score = score
    if best_score < 2:
        return None, []
    return best_row, best_values


def _profile_fields(
    table_name: str,
    table: TableSchema,
    schema_fields: list[str],
    dictionary_index: dict[tuple[str, str], dict[str, Any]],
) -> list[str]:
    fields: list[str] = []
    dictionary_fields = [field for current_table, field in dictionary_index if current_table == table_name]
    for field in [*table.primary_key, table.group_key or "", *dictionary_fields, *schema_fields]:
        if not field:
            continue
        if (
            field in table.primary_key
            or field == table.group_key
            or (table_name, field) in dictionary_index
            or _looks_like_context_field(field)
        ):
            fields.append(field)
        if len(fields) >= PROFILE_FIELD_LIMIT:
            break
    return _ordered_unique(fields)


def _empty_field_stat(
    field: str,
    table_name: str,
    table: TableSchema,
    dictionary_index: dict[tuple[str, str], dict[str, Any]],
) -> dict[str, Any]:
    dictionary_entry = dictionary_index.get((table_name, field), {})
    roles = _field_roles(field, table, dictionary_entry)
    allocation_role = _allocation_role(field, table, dictionary_entry, roles)
    return {
        "field": field,
        "roles": roles,
        "allocation_role": allocation_role,
        "id_strategy": dictionary_entry.get("id_strategy") or "",
        "reference_table": dictionary_entry.get("reference_table") or "",
        "risk_note": dictionary_entry.get("risk_note") or "",
        "non_empty_count": 0,
        "max_numeric": None,
        "max_numeric_row": None,
        "min_numeric": None,
        "last_numeric": None,
        "last_numeric_row": None,
        "last_value": None,
        "last_value_row": None,
        "_sample_values": [],
        "_distinct_values": set(),
        "_top_values": Counter(),
    }


def _field_roles(field: str, table: TableSchema, dictionary_entry: dict[str, Any]) -> list[str]:
    roles: list[str] = []
    lower = field.lower()
    strategy = str(dictionary_entry.get("id_strategy") or "").strip()
    if field in table.primary_key:
        roles.append("primary_key")
    if field == table.group_key or _looks_like_group_field(field):
        roles.append("group_key")
    if strategy == "lookup" or dictionary_entry.get("reference_table"):
        roles.append("lookup_ref")
    if "id" in lower or "ID" in field or "编号" in field:
        roles.append("id_like")
    if any(keyword in lower for keyword in ("list", "form")) or "列表" in field:
        roles.append("list_or_form")
    if _looks_like_enum_field(field):
        roles.append("enum_hint")
    if _looks_like_reference_field(field) and field not in table.primary_key and field != table.group_key:
        roles.append("foreign_key_candidate")
    return _ordered_unique(roles)


def _allocation_role(field: str, table: TableSchema, dictionary_entry: dict[str, Any], roles: list[str]) -> str | None:
    strategy = str(dictionary_entry.get("id_strategy") or "").strip()
    writable = dictionary_entry.get("writable", True)
    if writable is False or strategy == "lookup":
        return None
    if field in table.primary_key:
        return "primary_key"
    if field == table.group_key or ("group_key" in roles and strategy in {"", "unknown", *GENERATED_STRATEGIES}):
        return "group_key"
    if strategy in GENERATED_STRATEGIES:
        return "field_dictionary"
    return None


def _update_field_stat(stat: dict[str, Any], value: Any, row_index: int) -> None:
    stat["non_empty_count"] += 1
    stat["last_value"] = value
    stat["last_value_row"] = row_index
    value_key = _value_key(value)
    stat["_distinct_values"].add(value_key)
    stat["_top_values"][value_key] += 1
    if len(stat["_sample_values"]) < SAMPLE_VALUE_LIMIT and value_key not in {_value_key(item) for item in stat["_sample_values"]}:
        stat["_sample_values"].append(value)
    numeric = _numeric_value(value)
    if numeric is None:
        return
    stat["last_numeric"] = numeric
    stat["last_numeric_row"] = row_index
    if stat["max_numeric"] is None or numeric > stat["max_numeric"]:
        stat["max_numeric"] = numeric
        stat["max_numeric_row"] = row_index
    if stat["min_numeric"] is None or numeric < stat["min_numeric"]:
        stat["min_numeric"] = numeric


def _finalize_field_stat(stat: dict[str, Any], scanned_rows: int) -> dict[str, Any]:
    distinct_values = stat.pop("_distinct_values")
    top_values = stat.pop("_top_values")
    sample_values = stat.pop("_sample_values")
    non_empty = stat["non_empty_count"]
    fill_rate = round(non_empty / scanned_rows, 4) if scanned_rows else 0
    stat["fill_rate"] = fill_rate
    stat["distinct_count"] = len(distinct_values)
    stat["sample_values"] = sample_values
    stat["top_values"] = [{"value": value, "count": count} for value, count in top_values.most_common(SAMPLE_VALUE_LIMIT)]
    if stat["max_numeric"] is not None:
        stat["next_value"] = stat["max_numeric"] + 1
    else:
        stat["next_value"] = None
    if 0 < len(distinct_values) <= ENUM_VALUE_LIMIT and (_looks_like_enum_field(stat["field"]) or len(distinct_values) <= 5):
        stat["enum_values"] = [item["value"] for item in stat["top_values"]]
    else:
        stat["enum_values"] = []
    return stat


def _looks_like_context_field(field: str) -> bool:
    lower = field.lower()
    return _looks_like_reference_field(field) or _looks_like_enum_field(field) or any(keyword in lower for keyword in REFERENCE_KEYWORDS)


def _looks_like_reference_field(field: str) -> bool:
    lower = field.lower()
    return any(keyword in lower for keyword in REFERENCE_KEYWORDS) or any(word in field for word in REFERENCE_WORDS)


def _looks_like_group_field(field: str) -> bool:
    lower = field.lower()
    return any(word in lower for word in GROUP_WORDS) or "组" in field


def _looks_like_enum_field(field: str) -> bool:
    lower = field.lower()
    return any(hint in lower for hint in ENUM_HINTS) or any(hint in field for hint in ("类型", "状态", "开关", "是否"))


def _numeric_value(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    if text.isdigit():
        return int(text)
    return None


def _cell_value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _value_key(value: Any) -> str:
    return "" if value is None else str(value)


def _dictionary_index(entries: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    result: dict[tuple[str, str], dict[str, Any]] = {}
    for entry in entries:
        if entry.get("enabled") is False:
            continue
        table = str(entry.get("target_table") or "").strip()
        field = str(entry.get("target_field") or "").strip()
        if table and field:
            result[(table, field)] = entry
    return result


def _ordered_unique(values: list[Any]) -> list[Any]:
    result = []
    seen = set()
    for value in values:
        if value in (None, "") or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result
