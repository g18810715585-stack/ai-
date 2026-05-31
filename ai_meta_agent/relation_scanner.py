from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config_discovery import discover_config_tables
from .io_utils import read_json, write_json, write_text
from .models import ConfigTableRef, Manifest, SchemaBundle, TableSchema, resolve_path


REFERENCE_WORDS = (
    "id",
    "group",
    "reward",
    "goods",
    "key",
    "list",
    "form",
    "task",
    "activity",
    "shop",
    "item",
    "mail",
    "jump",
    "name",
    "title",
    "desc",
    "奖励",
    "商品",
    "礼包",
    "活动",
    "任务",
    "组",
    "文案",
    "标题",
    "描述",
    "名称",
    "跳转",
    "商店",
    "道具",
    "邮件",
)

ENUM_WORDS = (
    "type",
    "status",
    "state",
    "flag",
    "open",
    "enable",
    "quality",
    "sort",
    "order",
    "count",
    "num",
    "price",
    "time",
    "date",
    "level",
    "rank",
    "weight",
    "类型",
    "状态",
    "开关",
    "是否",
    "数量",
    "价格",
    "时间",
    "等级",
    "排序",
    "权重",
)

TABLE_HINTS = {
    "reward": ("reward", "奖励", "发奖"),
    "goods": ("goods", "商品", "道具"),
    "active_shop": ("active_shop", "shop", "商店", "兑换"),
    "exchange": ("exchange", "礼包", "付费"),
    "activity": ("activity", "活动"),
    "activity_task_target": ("task", "任务"),
    "key": ("key", "文案", "标题", "描述", "名称", "文本"),
    "jump": ("jump", "跳转"),
    "mail": ("mail", "邮件"),
}

GENERIC_TABLE_PARTS = {"activity", "reward", "goods", "shop", "task", "item", "config"}

ID_INDEX_WORDS = (
    "id",
    "key",
    "编号",
    "索引",
    "唯一",
    "文案",
)

NON_REFERENCE_TOKENS = {"int", "str", "string", "float", "double", "bool", "boolean", "date", "datetime", "array", "any"}


@dataclass
class KeyIndex:
    table: str
    field: str
    kind: str
    values: set[str]
    samples: dict[str, list[int]]


@dataclass
class TableRows:
    table: str
    fields: list[str]
    rows: list[dict[str, Any]]
    path: str | None = None
    sheet: str | None = None


def _now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _blank(value: Any) -> bool:
    return value is None or value == ""


def _canon(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    text = str(value).strip()
    if not text:
        return None
    if text.lower() in NON_REFERENCE_TOKENS:
        return None
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer() and re.fullmatch(r"-?\d+(?:\.0+)?", text):
        return str(int(number))
    return text


def split_reference_values(value: Any) -> list[str]:
    if _blank(value):
        return []
    if isinstance(value, (list, tuple, set)):
        tokens: list[str] = []
        for item in value:
            tokens.extend(split_reference_values(item))
        return list(dict.fromkeys(tokens))
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        canon = _canon(value)
        return [canon] if canon else []

    text = str(value).strip()
    if not text:
        return []
    if text.startswith("[") or text.startswith("{"):
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            parsed = None
        if parsed is not None:
            return split_reference_values(list(parsed.values()) if isinstance(parsed, dict) else parsed)

    parts = re.split(r"[\s,，;；|/]+", text)
    tokens = [_canon(part.strip("'\"[]{}()")) for part in parts if part.strip("'\"[]{}()")]
    tokens = [token for token in tokens if token]
    if len(tokens) <= 1:
        numbers = re.findall(r"(?<!\d)-?\d+(?!\d)", text)
        if numbers and not re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", text):
            tokens.extend(numbers)
    return list(dict.fromkeys(tokens))


def _field_text(field: str) -> str:
    return str(field or "").strip().lower()


def _is_config_table_name(name: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", str(name or "")))


def _has_reference_signal(field: str) -> bool:
    text = _field_text(field)
    return any(word in text for word in REFERENCE_WORDS)


def _is_enum_like(field: str) -> bool:
    text = _field_text(field)
    if text.startswith("is_") or text.startswith("has_") or "是否" in text:
        return True
    if any(word in text for word in ENUM_WORDS) and not any(word in text for word in ("id", "key", "list", "form", "文案")):
        return True
    if any(word in text for word in ("reward", "goods", "key", "list", "form", "奖励", "商品", "文案")):
        return False
    return any(word in text for word in ENUM_WORDS)


def _key_kind(field: str, primary_key: list[str]) -> str | None:
    text = _field_text(field)
    if field in primary_key:
        return "primary_key"
    if text == "group" or text.endswith("_group") or "组" in text:
        return "group_key"
    if text in {"key", "lang_key"} or "文案" in text:
        return "text_key"
    if any(word in text for word in ID_INDEX_WORDS) and not _is_enum_like(field):
        return "id_key"
    return None


def _candidate_source_fields(table: TableSchema) -> list[str]:
    fields: list[str] = []
    for field in table.fields:
        if field in table.primary_key:
            continue
        if not _has_reference_signal(field):
            continue
        if _is_enum_like(field):
            continue
        fields.append(field)
    return fields


def _candidate_key_fields(table: TableSchema) -> list[str]:
    fields: list[str] = []
    for field in table.primary_key:
        if field in table.fields and field not in fields:
            fields.append(field)
    for field in table.fields:
        if _key_kind(field, table.primary_key) and field not in fields:
            fields.append(field)
    return fields


def _hinted_target_tables(fields: list[str], schema: SchemaBundle) -> list[str]:
    hinted: list[str] = []
    for field in fields:
        text = _field_text(field)
        for table_name in schema.tables:
            if not _is_config_table_name(table_name):
                continue
            if table_name in hinted:
                continue
            if table_name.lower() in text:
                hinted.append(table_name)
                continue
            hints = TABLE_HINTS.get(table_name, ())
            if any(hint in text for hint in hints):
                hinted.append(table_name)
    return hinted


def _load_common_table_names(base_dir: Path) -> list[str]:
    tier_rank = {"core": 0, "high": 1, "medium": 2, "low": 3}
    always_include = {"active_shop", "reward", "goods", "exchange", "key", "jump", "mail", "activity_task_target"}
    for path in [base_dir / ".knowledge" / "common-tables.json", base_dir / "config" / "common-tables.json"]:
        if not path.exists():
            continue
        data = read_json(path)
        raw = data if isinstance(data, list) else data.get("tables", [])
        names = []
        for item in raw:
            if isinstance(item, str):
                name = item
                tier = "high"
            else:
                name = item.get("name") or item.get("sheet") or item.get("key")
                tier = str(item.get("frequencyTier") or item.get("frequency_tier") or "low").lower()
            if name and (name in always_include or tier_rank.get(tier, 99) <= 1):
                names.append(str(name))
        return list(dict.fromkeys([*names, *always_include]))
    return []


def _schema_scan_for(schema_path: Path) -> Path | None:
    sibling = schema_path.with_name("schema-scan.json")
    return sibling if sibling.exists() else None


def _refs_from_report(report: dict[str, Any]) -> dict[str, ConfigTableRef]:
    refs: dict[str, ConfigTableRef] = {}
    for table_name, table in (report.get("tables") or {}).items():
        path = table.get("source_file")
        if path:
            refs[table_name] = ConfigTableRef(path=path, sheet=table.get("sheet") or table_name)
    return refs


def _report_tables(report: dict[str, Any] | None) -> dict[str, Any]:
    return (report or {}).get("tables") or {}


def _header_row_from_report(report: dict[str, Any] | None, table_name: str) -> int | None:
    value = _report_tables(report).get(table_name, {}).get("header_row")
    return int(value) if value else None


def _field_headers(sheet: Any, header_row: int | None, schema_fields: list[str]) -> tuple[int, dict[str, int]]:
    if header_row:
        rows_to_check = [header_row]
    else:
        rows_to_check = list(range(1, min(sheet.max_row or 1, 10) + 1))
    wanted = set(schema_fields)
    best_row = rows_to_check[0]
    best_map: dict[str, int] = {}
    best_score = -1
    for row in rows_to_check:
        column_map: dict[str, int] = {}
        score = 0
        for col in range(1, (sheet.max_column or 0) + 1):
            value = sheet.cell(row, col).value
            if value in (None, ""):
                continue
            name = str(value).strip()
            if name in wanted and name not in column_map:
                score += 1
                column_map[name] = col
        if score > best_score:
            best_score = score
            best_row = row
            best_map = column_map
    return best_row, best_map


def _load_table_rows(
    table_name: str,
    table: TableSchema,
    ref: ConfigTableRef,
    base_dir: Path,
    fields: list[str],
    header_row: int | None = None,
    max_rows: int = 3000,
) -> TableRows:
    path = resolve_path(base_dir, ref.path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = ref.sheet or table.sheet or table_name
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        header_row, column_map = _field_headers(sheet, header_row, list(table.fields.keys()))
        fields = [field for field in fields if field in column_map]
        rows: list[dict[str, Any]] = []
        max_row = min(sheet.max_row or header_row, header_row + max_rows)
        max_col = max((column_map[field] for field in fields), default=0)
        if not max_col:
            return TableRows(table=table_name, fields=[], rows=[], path=str(path), sheet=sheet.title)
        for row_idx, raw in enumerate(
            sheet.iter_rows(min_row=header_row + 1, max_row=max_row, min_col=1, max_col=max_col, values_only=True),
            start=header_row + 1,
        ):
            row = {"__row": row_idx}
            has_value = False
            for field in fields:
                value = raw[column_map[field] - 1] if column_map[field] - 1 < len(raw) else None
                if not _blank(value):
                    has_value = True
                row[field] = value
            if has_value:
                rows.append(row)
        return TableRows(table=table_name, fields=fields, rows=rows, path=str(path), sheet=sheet.title)
    finally:
        workbook.close()


def _build_index(table_name: str, table: TableSchema, rows: TableRows) -> list[KeyIndex]:
    indexes = []
    for field in rows.fields:
        kind = _key_kind(field, table.primary_key)
        if not kind:
            continue
        samples: dict[str, list[int]] = defaultdict(list)
        values: set[str] = set()
        for row in rows.rows:
            token = _canon(row.get(field))
            if not token:
                continue
            values.add(token)
            if len(samples[token]) < 3:
                samples[token].append(int(row["__row"]))
        if len(values) >= 1:
            indexes.append(KeyIndex(table=table_name, field=field, kind=kind, values=values, samples=dict(samples)))
    return indexes


def _field_target_bonus(field: str, target_table: str, target_field: str) -> float:
    text = _field_text(field)
    target = target_table.lower()
    bonus = 0.0
    if target in text:
        bonus += 0.28
    for part in target.split("_"):
        if len(part) >= 4 and part not in GENERIC_TABLE_PARTS and part in text:
            bonus += 0.08
    if _field_text(target_field) in text and target_field.lower() not in {"id", "key"}:
        bonus += 0.12
    for table, hints in TABLE_HINTS.items():
        if table == target_table and any(hint in text for hint in hints):
            bonus += 0.22
    if "group" in text and target_field.lower() == "group":
        bonus += 0.18
    if "组" in text and "组" in target_field:
        bonus += 0.18
    return min(bonus, 0.45)


def _relation_type(field: str, target_field: str, tokens_per_value: list[int]) -> str:
    text = _field_text(field)
    target = _field_text(target_field)
    if "key" in text or "文案" in text or "标题" in text or "描述" in text or target in {"key", "lang_key"}:
        return "text_key_ref"
    if any(count > 1 for count in tokens_per_value) or "list" in text or "form" in text:
        return "list_ref"
    if "group" in text or "组" in text or "group" in target or "组" in target:
        return "group_ref"
    return "foreign_key"


def _score_relation(field: str, index: KeyIndex, hit_rate: float, hit_count: int, distinct_hits: int) -> float:
    bonus = _field_target_bonus(field, index.table, index.field)
    count_score = min(distinct_hits, 20) / 100
    confidence = 0.28 + hit_rate * 0.45 + count_score + bonus
    if hit_count < 2:
        confidence -= 0.15
    return round(max(0.0, min(confidence, 0.98)), 3)


def _risk(confidence: float, hit_rate: float) -> str:
    if confidence >= 0.78 and hit_rate >= 0.7:
        return "low"
    if confidence >= 0.58:
        return "medium"
    return "high"


def _analyze_source_table(source: TableRows, indexes: list[KeyIndex], root_targets: set[str], hop: int) -> list[dict[str, Any]]:
    relations: list[dict[str, Any]] = []
    for field in source.fields:
        tokens: list[str] = []
        token_rows: dict[str, list[int]] = defaultdict(list)
        tokens_per_value: list[int] = []
        for row in source.rows:
            parsed = split_reference_values(row.get(field))
            tokens_per_value.append(len(parsed))
            for token in parsed:
                tokens.append(token)
                if len(token_rows[token]) < 3:
                    token_rows[token].append(int(row["__row"]))
        if not tokens:
            continue
        distinct = set(tokens)
        if len(distinct) < 2 and not _field_target_bonus(field, source.table, field):
            continue
        for index in indexes:
            if index.table == source.table:
                continue
            hits = [token for token in tokens if token in index.values]
            distinct_hits = sorted({token for token in hits})
            if not hits:
                continue
            hit_rate = len(hits) / len(tokens)
            bonus = _field_target_bonus(field, index.table, index.field)
            if bonus <= 0 and hit_rate < 0.85:
                continue
            if len(hits) < 2 and bonus < 0.25:
                continue
            confidence = _score_relation(field, index, hit_rate, len(hits), len(distinct_hits))
            if confidence < 0.45:
                continue
            relations.append(
                {
                    "from_table": source.table,
                    "from_field": field,
                    "to_table": index.table,
                    "to_field": index.field,
                    "to_field_kind": index.kind,
                    "relation_type": _relation_type(field, index.field, tokens_per_value),
                    "confidence": confidence,
                    "risk": _risk(confidence, hit_rate),
                    "hop": hop,
                    "is_from_target": source.table in root_targets,
                    "evidence": {
                        "source_value_count": len(tokens),
                        "distinct_source_values": len(distinct),
                        "hit_count": len(hits),
                        "distinct_hit_count": len(distinct_hits),
                        "hit_rate": round(hit_rate, 3),
                        "matched_values": distinct_hits[:10],
                        "source_rows": sorted({row for token in distinct_hits[:10] for row in token_rows[token]})[:10],
                    },
                    "notes": "deterministic value match",
                }
            )
    relations.sort(key=lambda item: (-item["confidence"], item["from_table"], item["from_field"], item["to_table"]))
    return relations


def _dedupe_relations(relations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    best: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for relation in relations:
        key = (relation["from_table"], relation["from_field"], relation["to_table"], relation["to_field"])
        if key not in best or relation["confidence"] > best[key]["confidence"]:
            best[key] = relation
    return sorted(best.values(), key=lambda item: (item["hop"], -item["confidence"], item["from_table"], item["from_field"], item["to_table"]))


def _relation_markdown(result: dict[str, Any]) -> str:
    lines = ["# Relationship Map", ""]
    summary = result["summary"]
    lines.append(
        f"- targets={', '.join(result['target_tables']) or '-'} relations={summary['relation_count']} recommended={len(result['recommended_tables'])}"
    )
    if result["recommended_tables"]:
        lines.append(f"- recommended: {', '.join(result['recommended_tables'])}")
    lines.append("")
    for relation in result["relations"][:80]:
        evidence = relation["evidence"]
        lines.append(
            f"- {relation['from_table']}.{relation['from_field']} -> {relation['to_table']}.{relation['to_field']} "
            f"{relation['relation_type']} confidence={relation['confidence']:.2f} hit_rate={evidence['hit_rate']:.2f}"
        )
    lines.append("")
    return "\n".join(lines)


def compact_relationship_context(result: dict[str, Any], limit: int = 80) -> dict[str, Any]:
    return {
        "version": result.get("version"),
        "target_tables": result.get("target_tables", []),
        "recommended_tables": result.get("recommended_tables", []),
        "summary": result.get("summary", {}),
        "relations": result.get("relations", [])[:limit],
        "ai_review": result.get("ai_review"),
    }


def scan_relationships(
    manifest: Manifest,
    schema: SchemaBundle,
    base_dir: Path,
    run_dir: Path,
    schema_scan_path: Path | None = None,
    max_rows: int = 1500,
) -> dict[str, Any]:
    report = read_json(schema_scan_path) if schema_scan_path and schema_scan_path.exists() else None
    refs = _refs_from_report(report or {})
    refs.update(manifest.config_tables)
    if not refs and manifest.config_roots:
        manifest, _ = discover_config_tables(manifest, schema, base_dir)
        refs.update(manifest.config_tables)

    root_targets = [name for name in (manifest.target_tables or list(schema.tables.keys())) if name in schema.tables]
    common_names = _load_common_table_names(base_dir)
    root_candidate_fields = [
        field
        for table_name in root_targets
        for field in _candidate_source_fields(schema.tables[table_name])
    ]
    if len(schema.tables) <= 120:
        extra_index_names = list(refs.keys())
    else:
        extra_index_names = [*common_names, *_hinted_target_tables(root_candidate_fields, schema)]
    index_table_names = [
        name
        for name in dict.fromkeys([*root_targets, *extra_index_names])
        if _is_config_table_name(name) and name in schema.tables and name in refs
    ]

    diagnostics: dict[str, Any] = {"errors": [], "missing_refs": [], "indexed_tables": [], "analyzed_tables": []}
    key_indexes: list[KeyIndex] = []
    table_cache: dict[tuple[str, tuple[str, ...]], TableRows] = {}

    def load_rows(table_name: str, fields: list[str]) -> TableRows | None:
        key = (table_name, tuple(fields))
        if key in table_cache:
            return table_cache[key]
        ref = refs.get(table_name)
        if not ref:
            diagnostics["missing_refs"].append(table_name)
            return None
        try:
            rows = _load_table_rows(
                table_name,
                schema.tables[table_name],
                ref,
                base_dir,
                fields,
                _header_row_from_report(report, table_name),
                max_rows,
            )
        except Exception as exc:  # noqa: BLE001 - keep the rest of the relation scan useful.
            diagnostics["errors"].append({"table": table_name, "message": str(exc)})
            return None
        table_cache[key] = rows
        return rows

    for table_name in index_table_names:
        key_fields = _candidate_key_fields(schema.tables[table_name])
        rows = load_rows(table_name, key_fields)
        if not rows:
            continue
        key_indexes.extend(_build_index(table_name, schema.tables[table_name], rows))
        diagnostics["indexed_tables"].append(table_name)

    source_tables = [name for name in root_targets if name in refs]
    first_pass: list[dict[str, Any]] = []
    for table_name in source_tables:
        fields = _candidate_source_fields(schema.tables[table_name])
        rows = load_rows(table_name, fields)
        if not rows:
            continue
        diagnostics["analyzed_tables"].append(table_name)
        first_pass.extend(_analyze_source_table(rows, key_indexes, set(root_targets), 1))

    one_hop = sorted(
        {
            relation["to_table"]
            for relation in first_pass
            if relation["confidence"] >= 0.6 and relation["to_table"] not in root_targets and relation["to_table"] in refs
        }
    )
    second_pass: list[dict[str, Any]] = []
    for table_name in one_hop:
        fields = _candidate_source_fields(schema.tables[table_name])
        rows = load_rows(table_name, fields)
        if not rows:
            continue
        diagnostics["analyzed_tables"].append(table_name)
        second_pass.extend(_analyze_source_table(rows, key_indexes, set(root_targets), 2))

    relations = _dedupe_relations([*first_pass, *second_pass])
    high_confidence = [item for item in relations if item["confidence"] >= 0.75]
    medium_confidence = [item for item in relations if 0.55 <= item["confidence"] < 0.75]
    recommended = sorted(
        {
            item["to_table"]
            for item in relations
            if item["confidence"] >= (0.6 if item["hop"] == 1 else 0.75) and item["to_table"] not in root_targets
        }
    )

    result = {
        "version": 1,
        "generated_at": _now(),
        "project": manifest.project,
        "target_tables": root_targets,
        "analyzed_tables": sorted(set(diagnostics["analyzed_tables"])),
        "recommended_tables": recommended,
        "summary": {
            "relation_count": len(relations),
            "high_confidence_count": len(high_confidence),
            "medium_confidence_count": len(medium_confidence),
            "indexed_table_count": len(set(diagnostics["indexed_tables"])),
            "error_count": len(diagnostics["errors"]),
        },
        "relations": relations,
        "ai_review": {
            "mode": "deterministic_context",
            "summary": "已先用本地算法生成关系证据；真实 AI 生成草案时会读取这份关系图进行解释和补全。",
            "recommended_tables": recommended,
            "needs_confirmation": [item for item in relations if item["confidence"] < 0.65][:20],
        },
        "diagnostics": diagnostics,
    }
    write_json(run_dir / "relationship-map.json", result)
    write_text(run_dir / "relationship-map.md", _relation_markdown(result))
    return result
