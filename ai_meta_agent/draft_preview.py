from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import Manifest, Patch, PatchOperation, SchemaBundle


def build_draft_table_preview(
    manifest: Manifest,
    schema: SchemaBundle,
    patch: Patch,
    base_dir: Path,
    max_changed_rows_per_table: int = 120,
) -> dict[str, Any]:
    """Build a compact table-shaped view for reviewing a draft patch.

    The UI only needs the field area and rows touched by the draft, so this
    intentionally avoids loading or serializing entire configuration sheets.
    """
    operations_by_table: dict[str, list[PatchOperation]] = defaultdict(list)
    for operation in patch.operations:
        operations_by_table[operation.target_table].append(operation)

    tables = []
    for table_name, operations in operations_by_table.items():
        table_schema = schema.tables.get(table_name)
        config_ref = manifest.config_tables.get(table_name)
        fields = _field_order(table_schema, operations)
        source_path = _resolve_config_path(base_dir, config_ref.path if config_ref else None)
        sheet_name = (config_ref.sheet if config_ref else None) or (table_schema.sheet if table_schema else None) or table_name
        sheet_snapshot = _read_sheet_snapshot(source_path, sheet_name, fields)
        fields = _ordered_unique([*fields, *sheet_snapshot.get("detected_fields", [])])
        changed_rows = _changed_rows_for_table(
            operations,
            fields,
            sheet_snapshot.get("data_rows", []),
            max_changed_rows_per_table,
        )
        tables.append(
            {
                "table": table_name,
                "sheet": sheet_name,
                "source_file": str(source_path) if source_path else "",
                "fields": fields,
                "header_rows": _header_rows(sheet_snapshot.get("first_rows", []), fields),
                "changed_rows": changed_rows,
                "operation_count": len(operations),
                "changed_row_count": len(changed_rows),
                "warnings": sheet_snapshot.get("warnings", []),
            }
        )

    return {
        "patch_id": patch.patch_id,
        "project": patch.project,
        "table_count": len(tables),
        "tables": sorted(tables, key=lambda item: item["table"]),
    }


def _field_order(table_schema: Any, operations: list[PatchOperation]) -> list[str]:
    fields = list((table_schema.fields if table_schema else {}).keys())
    for operation in operations:
        fields.extend(operation.match.keys())
        fields.extend(operation.set.keys())
        for row in operation.rows:
            fields.extend(row.keys())
    return _ordered_unique([str(field) for field in fields if str(field or "").strip()])


def _resolve_config_path(base_dir: Path, value: str | None) -> Path | None:
    if not value:
        return None
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def _read_sheet_snapshot(source_path: Path | None, sheet_name: str, fields: list[str]) -> dict[str, Any]:
    if not source_path or not source_path.exists():
        return {"first_rows": [], "data_rows": [], "detected_fields": [], "warnings": ["未找到原始配置表文件，已仅按草案字段展示。"]}
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surfaced as a preview warning.
        return {"first_rows": [], "data_rows": [], "detected_fields": [], "warnings": [f"读取配置表失败：{exc}"]}
    try:
        if sheet_name not in workbook.sheetnames:
            return {"first_rows": [], "data_rows": [], "detected_fields": [], "warnings": [f"文件中没有 sheet：{sheet_name}"]}
        sheet = workbook[sheet_name]
        first_rows: list[list[Any]] = []
        raw_rows: list[list[Any]] = []
        max_cols = max(len(fields), sheet.max_column or 0)
        for index, row in enumerate(sheet.iter_rows(max_row=min(sheet.max_row or 0, 5000), max_col=max_cols, values_only=True), start=1):
            values = [_cell_value(value) for value in row]
            if index <= 3:
                first_rows.append(values)
            raw_rows.append(values)
        field_row_index, detected_fields = _detect_field_row(raw_rows[:10], fields)
        data_rows = _data_rows(raw_rows, field_row_index, detected_fields)
        return {"first_rows": first_rows, "data_rows": data_rows, "detected_fields": detected_fields, "warnings": []}
    finally:
        workbook.close()


def _detect_field_row(rows: list[list[Any]], expected_fields: list[str]) -> tuple[int | None, list[str]]:
    expected = {str(field) for field in expected_fields if str(field or "").strip()}
    best_index: int | None = None
    best_score = -1
    best_values: list[str] = []
    for index, row in enumerate(rows):
        values = [str(value or "").strip() for value in row]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue
        overlap = len(set(non_empty) & expected)
        score = overlap * 4 + len(non_empty)
        if score > best_score:
            best_index = index
            best_score = score
            best_values = values
    if best_index is None:
        return None, expected_fields
    detected = [value for value in best_values if value]
    return best_index, detected or expected_fields


def _data_rows(rows: list[list[Any]], field_row_index: int | None, fields: list[str]) -> list[dict[str, Any]]:
    if field_row_index is None or not fields:
        return []
    result = []
    for raw_index, raw in enumerate(rows[field_row_index + 1 :], start=field_row_index + 2):
        row = {field: raw[idx] if idx < len(raw) else None for idx, field in enumerate(fields)}
        if any(value not in (None, "") for value in row.values()):
            row["__row_number"] = raw_index
            result.append(row)
    return result


def _changed_rows_for_table(
    operations: list[PatchOperation],
    fields: list[str],
    data_rows: list[dict[str, Any]],
    max_rows: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for operation_index, operation in enumerate(operations, start=1):
        if len(rows) >= max_rows:
            break
        if operation.op == "insert":
            for row in operation.rows:
                rows.append(_preview_row("新增", operation, operation_index, fields, row, before=None))
                if len(rows) >= max_rows:
                    break
        elif operation.op == "update":
            matches = _find_matching_rows(data_rows, operation.match)
            if not matches:
                after = {**operation.match, **operation.set}
                rows.append(_preview_row("修改", operation, operation_index, fields, after, before=None))
                continue
            for before in matches:
                after = {key: value for key, value in before.items() if key != "__row_number"}
                after.update(operation.set)
                rows.append(_preview_row("修改", operation, operation_index, fields, after, before=before))
                if len(rows) >= max_rows:
                    break
        elif operation.op == "replace_group":
            for row in operation.rows:
                rows.append(_preview_row("替换后新增", operation, operation_index, fields, row, before=None))
                if len(rows) >= max_rows:
                    break
        elif operation.op == "delete_where":
            for before in _find_matching_rows(data_rows, operation.match)[: max_rows - len(rows)]:
                rows.append(_preview_row("删除匹配", operation, operation_index, fields, before, before=before))
    return rows[:max_rows]


def _preview_row(
    row_kind: str,
    operation: PatchOperation,
    operation_index: int,
    fields: list[str],
    values: dict[str, Any],
    before: dict[str, Any] | None,
) -> dict[str, Any]:
    row_number = values.get("__row_number") if isinstance(values, dict) else None
    clean_before = {field: before.get(field) for field in fields} if before else None
    return {
        "row_kind": row_kind,
        "op": operation.op,
        "operation_index": operation_index,
        "row_number": row_number,
        "values": {field: values.get(field) for field in fields},
        "before": clean_before,
        "changed_fields": _changed_fields(operation, fields),
        "match": operation.match,
        "reason": operation.reason,
        "confidence": operation.confidence,
        "risk_level": operation.risk_level,
        "needs_confirmation": operation.needs_confirmation,
    }


def _changed_fields(operation: PatchOperation, fields: list[str]) -> list[str]:
    if operation.op == "update":
        return [field for field in operation.set if field in fields]
    if operation.op in {"insert", "replace_group"}:
        return [field for field in fields if any(field in row for row in operation.rows)]
    return []


def _find_matching_rows(data_rows: list[dict[str, Any]], match: dict[str, Any]) -> list[dict[str, Any]]:
    if not match:
        return []
    result = []
    for row in data_rows:
        if all(_same_value(row.get(field), value) for field, value in match.items()):
            result.append(row)
    return result


def _same_value(left: Any, right: Any) -> bool:
    return str(left if left is not None else "").strip() == str(right if right is not None else "").strip()


def _header_rows(first_rows: list[list[Any]], fields: list[str]) -> list[dict[str, Any]]:
    rows = []
    for index in range(3):
        raw = first_rows[index] if index < len(first_rows) else []
        rows.append(
            {
                "row_kind": f"字段区第 {index + 1} 行",
                "row_number": index + 1,
                "values": {field: raw[field_index] if field_index < len(raw) else None for field_index, field in enumerate(fields)},
            }
        )
    return rows


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def _ordered_unique(values: list[Any]) -> list[Any]:
    result = []
    seen = set()
    for value in values:
        if value in (None, "") or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result
