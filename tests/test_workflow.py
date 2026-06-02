from __future__ import annotations

import json
import os
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from typing import Any
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ai_meta_agent.ai_context import build_minimal_context
from ai_meta_agent.cli import _auto_expand_generation_tables, analyze_manifest
from ai_meta_agent.configuration_records import build_configuration_record, local_case_review, save_case_review
from ai_meta_agent.context_optimizer import build_context_budget, compact_target_table_profiles, enforce_fast_context_budget, optimize_context_for_ai
from ai_meta_agent.draft import call_baseai, call_draft_diagnostics_ai, call_experience_summary_ai, call_relationship_ai, make_stub_patch
from ai_meta_agent.draft_diagnostics import build_draft_diagnostics, compact_draft_diagnostic_context
from ai_meta_agent.draft_preview import build_draft_table_preview
from ai_meta_agent.experience import (
    append_case_from_patch,
    build_structured_correction,
    build_experience_context,
    delete_saved_experience,
    list_activity_templates,
    list_field_dictionary,
    list_saved_experiences,
    merge_experience_summary,
    save_structured_correction,
    seed_field_dictionary_from_schema,
    summarize_experience_locally,
    teach_experience,
    upsert_field_dictionary_entry,
    update_saved_experience,
)
from ai_meta_agent.feishu import FeishuSourcePayload, read_feishu_sheet
from ai_meta_agent.habits import append_habit, habit_from_patch, load_habits, match_habits
from ai_meta_agent.id_allocator import fill_active_shop_incremental_ids, fill_incremental_placeholders
from ai_meta_agent.io_utils import read_json, write_json
from ai_meta_agent.item_resolution import resolve_planning_items
from ai_meta_agent.models import Manifest, Patch, PlanningSource, SchemaBundle, SheetIR, SourceKind, WorkbookIR
from ai_meta_agent.patch_sanitizer import sanitize_patch
from ai_meta_agent.patch_engine import apply_patch
from ai_meta_agent.relation_scanner import scan_relationships, split_reference_values
from ai_meta_agent.schema import load_schema
from ai_meta_agent.schema_scanner import scan_config_schema
from ai_meta_agent.table_profiles import build_target_table_profiles
from ai_meta_agent.workbook_ir import load_source_ir


ROOT = Path(__file__).resolve().parents[1]


def make_planning(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "礼包规划"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "六月礼包"
    headers = ["礼包ID", "礼包名称", "价格ID", "限购次数", "开始时间", "结束时间"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(2, col).value = header
    sheet.append([1001, "每日礼包", 68, None, "2026-06-01 05:00:00", "2026-06-08 04:59:59"])
    sheet.append([1002, "进阶礼包", 128, 1, "2026-06-01 05:00:00", "2026-06-08 04:59:59"])
    sheet["B4"].fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    sheet["F4"].comment = Comment("示例备注", "test")
    sheet.column_dimensions["G"].hidden = True
    sheet["G2"] = "隐藏列"
    workbook.save(path)


def make_config(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "shop_pack_config"
    headers = [
        "pack_id",
        "name",
        "price_id",
        "limit_type",
        "limit_count",
        "start_time",
        "end_time",
        "creator",
        "create_time",
        "internal_note",
    ]
    sheet.append(headers)
    sheet.append([1001, "旧每日礼包", 30, "daily", 1, "old-start", "old-end", "gaoyang", "2026-05-01", "keep"])
    reward = workbook.create_sheet("reward_item_config")
    reward.append(["reward_group_id", "item_id", "item_count", "item_order"])
    reward.append([1001, 3001, 10, 1])
    workbook.save(path)


def make_relation_config(path: Path) -> None:
    workbook = Workbook()
    activity = workbook.active
    activity.title = "activity"
    activity.append(["id", "form_list", "name_key"])
    activity.append([1, "100|101", "activity_title_1"])
    activity.append([2, "102", "activity_title_2"])

    shop = workbook.create_sheet("active_shop")
    shop.append(["id", "group", "goods", "title_key"])
    shop.append([10, 100, 500, "shop_title_1"])
    shop.append([11, 101, 501, "shop_title_2"])
    shop.append([12, 102, 502, "shop_title_3"])

    reward = workbook.create_sheet("reward")
    reward.append(["id", "type", "num"])
    reward.append([500, 1, 10])
    reward.append([501, 1, 20])
    reward.append([502, 1, 30])

    key = workbook.create_sheet("key")
    key.append(["key", "text"])
    key.append(["activity_title_1", "Activity 1"])
    key.append(["activity_title_2", "Activity 2"])
    key.append(["shop_title_1", "Shop 1"])
    key.append(["shop_title_2", "Shop 2"])
    key.append(["shop_title_3", "Shop 3"])
    workbook.save(path)


def make_manifest(tmp: Path, planning: Path, config: Path) -> Path:
    manifest = {
        "project": "unit-sample",
        "mode": "supervised_write",
        "schema_path": str(ROOT / "config" / "example.schema.json"),
        "run_root": str(tmp / ".runs"),
        "planning_sources": [{"id": "plan", "kind": "local_excel", "path": str(planning), "role": "planning"}],
        "config_tables": {
            "shop_pack_config": {"path": str(config), "sheet": "shop_pack_config"},
            "reward_item_config": {"path": str(config), "sheet": "reward_item_config"},
        },
        "habit_store": str(tmp / ".knowledge" / "habits.jsonl"),
    }
    path = tmp / "manifest.json"
    write_json(path, manifest)
    return path


class WorkflowTests(unittest.TestCase):
    def test_workbook_ir_and_stub_patch_flow(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            planning = tmp / "planning.xlsx"
            config = tmp / "config.xlsx"
            make_planning(planning)
            make_config(config)
            manifest_path = make_manifest(tmp, planning, config)

            manifest, schema, run_dir, context = analyze_manifest(manifest_path, tmp, "analysis")
            self.assertTrue((run_dir / "analysis.json").exists())
            sheet = context["workbooks"][0]["sheets"][0]
            self.assertEqual(sheet["header_row"], 2)
            self.assertIn("A1:F1", sheet["merged_ranges"])
            self.assertIn("G", sheet["hidden_columns"])

            patch = make_stub_patch(manifest, schema, context, str(tmp))
            self.assertEqual(len(patch.operations), 2)
            self.assertEqual([op.op for op in patch.operations], ["update", "insert"])
            self.assertTrue(all(op.needs_confirmation for op in patch.operations))
            draft_preview = build_draft_table_preview(manifest, schema, patch, tmp)
            self.assertEqual(draft_preview["table_count"], 1)
            table_preview = draft_preview["tables"][0]
            self.assertEqual(table_preview["table"], "shop_pack_config")
            self.assertEqual(len(table_preview["header_rows"]), 3)
            self.assertIn("pack_id", table_preview["fields"])
            self.assertEqual([row["row_kind"] for row in table_preview["changed_rows"]], ["修改", "新增"])
            changed_name = table_preview["changed_rows"][0]["values"]["name"]
            self.assertEqual(changed_name, "每日礼包")
            self.assertEqual(table_preview["changed_rows"][0]["before"]["name"], "旧每日礼包")

            apply_dir = tmp / ".runs" / "apply-test"
            apply_dir.mkdir(parents=True)
            result = apply_patch(manifest, schema, patch, tmp, apply_dir)
            self.assertTrue(result["previews"])
            preview = Path(next(iter(result["previews"].values())))
            self.assertTrue(preview.exists())
            workbook = load_workbook(preview, data_only=True)
            sheet = workbook["shop_pack_config"]
            rows = list(sheet.iter_rows(values_only=True))
            self.assertEqual(rows[1][1], "每日礼包")
            self.assertEqual(rows[1][7], "gaoyang")
            self.assertEqual(rows[2][0], 1002)
            validation = next(iter(result["validation"].values()))
            self.assertEqual(validation["errors"], [])
            self.assertTrue(result["rollback_patch"]["operations"])

    def test_apply_uses_second_header_row_for_machine_fields(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            reward = tmp / "reward.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "reward"
            display_headers = ["池id", "池名称", "奖励类型", "抽取类型", "抽取次数", "英雄限制", "等级限制", "大本系数", "是否触发借贷补偿", "备注", "类型", "内容", "数量"]
            field_headers = ["id", None, "type", "is_removal", "num", "hero_limit", "level_limit", "building_ratio", "is_repay", None, "type_1", "reward_1", "num_1"]
            type_headers = ["int", None, "int", "int", "int", "int", "int", "int", "int", None, "int", "string", "string"]
            for col, value in enumerate(display_headers, start=1):
                sheet.cell(1, col).value = value
            for col, value in enumerate(field_headers, start=1):
                sheet.cell(2, col).value = value
            for col, value in enumerate(type_headers, start=1):
                sheet.cell(3, col).value = value
            workbook.save(reward)

            manifest = Manifest.model_validate(
                {
                    "project": "reward-header-test",
                    "mode": "supervised_write",
                    "schema_path": "",
                    "planning_sources": [],
                    "config_tables": {"reward": {"path": str(reward), "sheet": "reward"}},
                }
            )
            schema = SchemaBundle.model_validate(
                {
                    "version": 1,
                    "tables": {
                        "reward": {
                            "sheet": "reward",
                            "primary_key": ["id"],
                            "fields": {
                                "id": {"type": "int", "required": True},
                                "type": {"type": "int"},
                                "is_removal": {"type": "int"},
                                "num": {"type": "int"},
                                "type_1": {"type": "int"},
                                "reward_1": {"type": "str"},
                                "num_1": {"type": "str"},
                            },
                        }
                    },
                }
            )
            patch_obj = Patch.model_validate(
                {
                    "patch_id": "reward-header-row",
                    "project": "reward-header-test",
                    "operations": [
                        {
                            "op": "insert",
                            "target_table": "reward",
                            "rows": [{"id": 605300001, "type": 4, "is_removal": 0, "num": 1, "type_1": 999, "reward_1": "10001", "num_1": "3"}],
                            "reason": "unit test",
                            "confidence": 1,
                            "risk_level": "low",
                            "needs_confirmation": False,
                        }
                    ],
                }
            )

            apply_dir = tmp / ".runs" / "apply-reward-header"
            result = apply_patch(manifest, schema, patch_obj, tmp, apply_dir)
            preview = Path(next(iter(result["previews"].values())))
            preview_sheet = load_workbook(preview, data_only=True)["reward"]
            self.assertEqual(preview_sheet.max_column, len(display_headers))
            self.assertEqual(preview_sheet.cell(4, 1).value, 605300001)
            self.assertEqual(preview_sheet.cell(4, 3).value, 4)
            self.assertEqual(preview_sheet.cell(4, 4).value, 0)
            self.assertEqual(preview_sheet.cell(4, 11).value, 999)
            self.assertEqual(preview_sheet.cell(4, 12).value, "10001")
            self.assertEqual(preview_sheet.cell(4, 13).value, "3")

    def test_inserted_rows_write_values_only_without_changing_grid_rules(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            styled = tmp / "styled.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "styled"
            sheet.sheet_view.showGridLines = False
            sheet.append(["编号", "名称"])
            sheet.append(["id", "name"])
            sheet.append(["int", "string"])
            thin = Side(style="thin", color="FF000000")
            existing_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row_idx in range(4, 7):
                sheet.cell(row_idx, 1).value = row_idx - 3
                sheet.cell(row_idx, 1).font = Font(name="Arial", sz=9)
                sheet.cell(row_idx, 1).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row_idx, 1).border = existing_border
                sheet.cell(row_idx, 2).value = f"item-{row_idx}"
                sheet.cell(row_idx, 2).font = Font(name="Courier New", sz=13)
                sheet.cell(row_idx, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                sheet.cell(row_idx, 2).border = existing_border
            sheet.cell(6, 1).font = Font(name="Times New Roman", sz=12)
            workbook.save(styled)

            manifest = Manifest.model_validate(
                {
                    "project": "style-test",
                    "mode": "supervised_write",
                    "schema_path": "",
                    "planning_sources": [],
                    "config_tables": {"styled": {"path": str(styled), "sheet": "styled"}},
                }
            )
            schema = SchemaBundle.model_validate(
                {
                    "version": 1,
                    "tables": {
                        "styled": {
                            "sheet": "styled",
                            "primary_key": ["id"],
                            "fields": {"id": {"type": "int"}, "name": {"type": "str"}},
                        }
                    },
                }
            )
            patch_obj = Patch.model_validate(
                {
                    "patch_id": "style-insert",
                    "project": "style-test",
                    "operations": [
                        {
                            "op": "insert",
                            "target_table": "styled",
                            "rows": [{"id": 4, "name": "new item"}],
                            "reason": "unit test",
                            "confidence": 1,
                        }
                    ],
                }
            )

            apply_dir = tmp / ".runs" / "apply-style"
            result = apply_patch(manifest, schema, patch_obj, tmp, apply_dir)
            preview = Path(next(iter(result["previews"].values())))
            preview_sheet = load_workbook(preview)["styled"]
            id_cell = preview_sheet.cell(7, 1)
            name_cell = preview_sheet.cell(7, 2)
            self.assertFalse(preview_sheet.sheet_view.showGridLines)
            self.assertEqual(id_cell.value, 4)
            self.assertEqual(name_cell.value, "new item")
            self.assertNotEqual(id_cell.font.name, "Arial")
            self.assertIsNone(id_cell.alignment.horizontal)
            self.assertNotEqual(name_cell.font.name, "Courier New")
            self.assertIsNone(name_cell.alignment.horizontal)
            self.assertIsNone(name_cell.alignment.wrap_text)
            self.assertEqual(preview_sheet.cell(4, 1).border.left.style, "thin")
            self.assertIsNone(id_cell.border.left.style)

    def test_habit_learning_round_trip(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            planning = tmp / "planning.xlsx"
            config = tmp / "config.xlsx"
            make_planning(planning)
            make_config(config)
            manifest_path = make_manifest(tmp, planning, config)
            manifest = Manifest.model_validate(json.loads(manifest_path.read_text(encoding="utf-8")))
            schema = load_schema(ROOT / "config" / "example.schema.json")
            _, _, _, context = analyze_manifest(manifest_path, tmp, "analysis")
            patch = make_stub_patch(manifest, schema, context, str(tmp))
            habit = habit_from_patch(patch, "accepted", "unit test accepted")
            habit_path = tmp / ".knowledge" / "habits.jsonl"
            append_habit(habit_path, habit)
            habits = load_habits(habit_path)
            self.assertEqual(len(habits), 1)
            matched = match_habits(habits, "unit-sample", ["shop_pack_config"])
            self.assertEqual(matched[0].habit_id, habit.habit_id)

    def test_overwrite_apply_writes_record_and_case_review(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            planning = tmp / "planning.xlsx"
            config = tmp / "config.xlsx"
            make_planning(planning)
            make_config(config)
            manifest_path = make_manifest(tmp, planning, config)
            manifest, schema, _, context = analyze_manifest(manifest_path, tmp, "analysis")
            patch = make_stub_patch(manifest, schema, context, str(tmp))

            apply_dir = tmp / ".runs" / "overwrite-test"
            apply_dir.mkdir(parents=True)
            result = apply_patch(manifest, schema, patch, tmp, apply_dir, write_mode="overwrite")
            self.assertEqual(result["write_mode"], "overwrite")
            self.assertTrue(result["backups"])
            self.assertTrue(result["written_files"])
            result["timing"] = {"total_seconds": 1.23, "apply_seconds": 0.9, "write_mode": "overwrite", "status": "success"}
            workbook = load_workbook(config, data_only=True)
            rows = list(workbook["shop_pack_config"].iter_rows(values_only=True))
            self.assertEqual(rows[2][0], 1002)

            record = build_configuration_record(manifest, patch, result, apply_dir)
            self.assertEqual(record["write_mode"], "overwrite")
            self.assertEqual(record["operation_count"], 2)
            self.assertIn("shop_pack_config", record["target_tables"])
            self.assertEqual(record["timing"]["total_seconds"], 1.23)
            review = local_case_review("价格字段要复核，礼包名称不能直接覆盖旧翻译。", record)
            case = save_case_review(tmp, manifest, patch, result, "价格字段要复核，礼包名称不能直接覆盖旧翻译。", review)
            self.assertEqual(case["decision"], "corrected")
            self.assertTrue((tmp / ".knowledge" / "case_examples.jsonl").exists())
            self.assertIn("价格字段", json.dumps(case, ensure_ascii=False))

    def test_patch_operation_accepts_string_source_ref_from_ai(self) -> None:
        patch_obj = Patch.model_validate(
            {
                "patch_id": "patch_string_ref",
                "project": "unit-sample",
                "operations": [
                    {
                        "op": "insert",
                        "target_table": "activity",
                        "rows": [{"id": 5805}],
                        "source_ref": "feishu-planning,2026年5月航海节活动,row 7",
                        "reason": "ai returned source_ref as text",
                        "confidence": 0.75,
                    }
                ],
            }
        )

        ref = patch_obj.operations[0].source_ref
        self.assertEqual(ref.workbook, "feishu-planning")
        self.assertEqual(ref.sheet, "2026年5月航海节活动")
        self.assertEqual(ref.row, 7)

    def test_patch_operation_accepts_common_ai_op_aliases(self) -> None:
        patch_obj = Patch.model_validate(
            {
                "patch_id": "patch_alias_ref",
                "project": "unit-sample",
                "operations": [
                    {
                        "op": "insert_rows",
                        "target_table": "activity",
                        "rows": [{"id": 5805}],
                        "source_ref": {"workbook": "plan"},
                        "reason": "ai returned insert_rows alias",
                        "confidence": 0.75,
                    }
                ],
            }
        )

        self.assertEqual(patch_obj.operations[0].op, "insert")

    def test_insert_operation_accepts_ai_set_payload_as_single_row(self) -> None:
        patch_obj = Patch.model_validate(
            {
                "patch_id": "patch_insert_set",
                "project": "unit-sample",
                "operations": [
                    {
                        "op": "insert",
                        "target_table": "activity",
                        "set": {"id": 5805, "name": "sample"},
                        "source_ref": {"workbook": "plan"},
                        "reason": "ai returned insert payload in set",
                        "confidence": 0.75,
                    }
                ],
            }
        )

        self.assertEqual(patch_obj.operations[0].rows, [{"id": 5805, "name": "sample"}])
        self.assertEqual(patch_obj.operations[0].set, {})

    def test_reward_unused_slots_are_removed_from_ai_patch(self) -> None:
        patch_obj = Patch.model_validate(
            {
                "patch_id": "patch_reward_slots",
                "project": "unit-sample",
                "operations": [
                    {
                        "op": "insert",
                        "target_table": "reward",
                        "rows": [
                            {
                                "id": 605300001,
                                "type_1": 7,
                                "reward_1": 323,
                                "num_1": 1,
                                "type_2": 0,
                                "reward_2": "0",
                                "num_2": 0,
                                "type_3": "",
                                "reward_3": None,
                                "num_3": "0",
                            },
                            {
                                "id": 605300002,
                                "type_1": 7,
                                "reward_1": 324,
                                "num_1": 1,
                                "type_2": 7,
                                "reward_2": 325,
                                "num_2": 1,
                            },
                        ],
                        "source_ref": {"workbook": "plan"},
                        "reason": "ai filled unused reward slots with placeholders",
                        "confidence": 0.75,
                    }
                ],
            }
        )

        result = sanitize_patch(patch_obj)
        first_row = patch_obj.operations[0].rows[0]
        second_row = patch_obj.operations[0].rows[1]
        self.assertNotIn("type_2", first_row)
        self.assertNotIn("reward_3", first_row)
        self.assertEqual(second_row["type_2"], 7)
        self.assertEqual(result["reward_unused_slots"]["removed_field_groups"], 2)

    def test_blank_insert_fields_are_removed_but_update_blanks_remain(self) -> None:
        patch_obj = Patch.model_validate(
            {
                "patch_id": "patch_blank_fields",
                "project": "unit-sample",
                "operations": [
                    {
                        "op": "insert",
                        "target_table": "activity",
                        "rows": [{"id": 1, "name": "activity", "empty": "", "unknown": None}],
                        "source_ref": {"workbook": "plan"},
                        "reason": "ai returned blank insert fields",
                        "confidence": 0.8,
                    },
                    {
                        "op": "update",
                        "target_table": "activity",
                        "match": {"id": 1},
                        "set": {"name": ""},
                        "source_ref": {"workbook": "plan"},
                        "reason": "explicit clear remains reviewable",
                        "confidence": 0.8,
                    },
                ],
            }
        )

        result = sanitize_patch(patch_obj)
        self.assertEqual(patch_obj.operations[0].rows[0], {"id": 1, "name": "activity"})
        self.assertEqual(patch_obj.operations[1].set, {"name": ""})
        self.assertEqual(result["blank_insert_fields"]["removed_fields"], 2)

    def test_teach_experience_writes_local_knowledge(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            result = teach_experience(
                tmp,
                "lesson-sample",
                "exchange shop activity usually uses activity, active_shop, exchange, reward, goods and key. "
                "planning item_id is a goods id, price maps to exchange price, reward maps to reward id.",
            )
            self.assertEqual(result["created"]["rules"], 1)
            self.assertEqual(result["created"]["activity_templates"], 0)
            self.assertGreaterEqual(result["created"]["field_mappings"], 2)
            self.assertTrue((tmp / ".knowledge" / "rules.jsonl").exists())
            self.assertEqual((tmp / ".knowledge" / "activity_templates.jsonl").read_text(encoding="utf-8").strip(), "")

    def test_saved_experience_history_update_and_delete(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            result = teach_experience(
                tmp,
                "history-sample",
                "兑换商店活动一般要看 activity、active_shop、exchange、reward、goods、key。商品名 -> goods.name。",
            )
            experience_id = result["experience_id"]
            listed = list_saved_experiences(tmp, project="history-sample")
            self.assertEqual(listed["count"], 1)
            self.assertEqual(listed["experiences"][0]["experience_id"], experience_id)
            self.assertTrue(listed["experiences"][0]["created_at"])

            updated = update_saved_experience(
                tmp,
                experience_id,
                "兑换商店活动一般要看 activity、exchange、reward。价格 -> exchange.price。",
                project="history-sample",
            )
            self.assertEqual(updated["experience"]["experience_id"], experience_id)
            self.assertIn("价格", updated["experience"]["text"])
            rules = [json.loads(line) for line in (tmp / ".knowledge" / "rules.jsonl").read_text(encoding="utf-8").splitlines()]
            self.assertEqual(len([item for item in rules if item.get("experience_id") == experience_id]), 1)

            deleted = delete_saved_experience(tmp, experience_id)
            self.assertEqual(deleted["deleted"], experience_id)
            self.assertEqual(list_saved_experiences(tmp, project="history-sample")["count"], 0)

    def test_experience_summary_is_reviewed_before_save(self) -> None:
        local = summarize_experience_locally(
            "lesson-sample",
            "兑换商店活动一般要看 activity、active_shop、exchange、reward、goods、key。商品名 -> goods.name，价格 -> exchange.price。",
        )
        self.assertEqual(local["mode"], "local")
        self.assertIn("字段映射", local["review_text"])
        self.assertGreaterEqual(len(local["records_preview"]["field_mappings"]), 2)

        merged = merge_experience_summary(
            "lesson-sample",
            "raw text",
            local,
            {
                "summary_title": "兑换商店字段经验",
                "review_text": "经验标题：兑换商店字段经验\n字段映射：\n- 商品名 -> goods.name\n- 价格 -> exchange.price",
                "field_mappings": [{"source_aliases": ["商品名"], "target_table": "goods", "target_field": "name"}],
                "activity_templates": [],
                "personal_rules": [],
                "questions": [],
                "risk_notes": ["保存前确认目标字段"],
                "conflicts": [],
            },
        )
        self.assertEqual(merged["mode"], "ai")
        self.assertIn("goods.name", merged["review_text"])
        self.assertTrue(merged["records_preview"]["field_mappings"])
        self.assertFalse(merged["has_conflicts"])

    def test_experience_summary_detects_mapping_conflicts(self) -> None:
        local = summarize_experience_locally(
            "lesson-sample",
            "兑换商店里商品名 -> reward.id",
            existing_experiences=[
                {
                    "experience_id": "exp_old",
                    "title": "旧商品映射",
                    "project": "lesson-sample",
                    "created_at": "2026-01-01T00:00:00Z",
                    "text": "兑换商店里商品名 -> goods.name",
                }
            ],
        )

        self.assertTrue(local["has_conflicts"])
        self.assertEqual(local["conflicts"][0]["conflict_type"], "field_mapping")
        self.assertEqual(local["conflicts"][0]["existing_experience_id"], "exp_old")

    def test_run_instruction_template_dictionary_and_context(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            manifest = Manifest.model_validate(
                {
                    "project": "bp-sample",
                    "mode": "supervised_write",
                    "schema_path": str(ROOT / "config" / "example.schema.json"),
                    "run_instruction": "本次是 BP 通行证活动，活动 ID 新建，奖励组按规划新建。",
                    "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx"}],
                    "target_tables": ["activity", "activity_task_target", "activity_point_mission", "reward"],
                }
            )
            schema = SchemaBundle.model_validate(
                {
                    "version": 1,
                    "tables": {
                        "activity": {"primary_key": ["id"], "fields": {"id": {"type": "int"}, "活动标题": {"type": "str"}}},
                        "activity_task_target": {"primary_key": ["id"], "fields": {"id": {"type": "int"}, "任务逻辑": {"type": "str"}, "目标值": {"type": "int"}}},
                        "activity_point_mission": {"primary_key": ["任务id"], "fields": {"任务id": {"type": "int"}, "奖励": {"type": "int"}, "积分": {"type": "int"}}},
                        "reward": {"primary_key": ["id"], "fields": {"id": {"type": "int"}}},
                    },
                }
            )
            workbook = WorkbookIR(
                source_id="plan",
                source_type=SourceKind.LOCAL_EXCEL,
                sheets=[
                    SheetIR(
                        name="BP规划",
                        max_row=3,
                        max_column=4,
                        headers=["活动ID", "任务", "奖励", "积分"],
                        header_row=1,
                        sample_rows=[{"活动ID": 9001, "任务": "每日登录", "奖励": 30001, "积分": 10}],
                    )
                ],
            )

            seed_field_dictionary_from_schema(tmp, schema)
            upsert_field_dictionary_entry(
                tmp,
                {
                    "target_table": "activity_point_mission",
                    "target_field": "奖励",
                    "description": "BP 任务奖励组",
                    "source_aliases": ["奖励", "奖励组"],
                    "id_strategy": "new",
                },
            )
            experience = build_experience_context(tmp, manifest, schema, [workbook], {})
            self.assertEqual(experience["config_plan"]["activity_template_id"], "battle_pass")
            self.assertTrue(experience["field_dictionary_matches"])
            self.assertEqual(experience["config_plan"]["run_instruction"], manifest.run_instruction)

            context = build_minimal_context(manifest, schema, [workbook], [], experience)
            self.assertEqual(context["run_instruction"], manifest.run_instruction)
            self.assertIn("field_dictionary_matches", context)

    def test_ai_context_keeps_more_planning_rows_for_corrections(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "context-row-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx"}],
            }
        )
        schema = SchemaBundle.model_validate(
            {
                "version": 1,
                "tables": {
                    "active_shop": {
                        "primary_key": ["id"],
                        "fields": {"id": {"type": "str"}, "商品": {"type": "str"}},
                    }
                },
            }
        )
        rows = [{"__row": index, "商品": f"商品{index}"} for index in range(1, 251)]
        workbook = WorkbookIR(
            source_id="feishu-planning",
            source_type=SourceKind.FEISHU,
            sheets=[
                SheetIR(
                    name="兑换店规划",
                    max_row=250,
                    max_column=2,
                    headers=["商品"],
                    header_row=1,
                    sample_rows=rows,
                )
            ],
        )

        context = build_minimal_context(manifest, schema, [workbook], [], None)
        sample_rows = context["workbooks"][0]["sheets"][0]["sample_rows"]
        self.assertIn(31, [row["__row"] for row in sample_rows])
        self.assertEqual(len(sample_rows), 200)
        self.assertEqual(context["workbooks"][0]["sheets"][0]["sample_row_count"], 250)
        self.assertEqual(context["workbooks"][0]["sheets"][0]["sample_rows_omitted"], 50)

    def test_ai_context_keeps_large_value_table_rows(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "context-value-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [{"id": "value", "kind": "local_excel", "path": "dummy.xlsx", "role": "item_base"}],
            }
        )
        schema = SchemaBundle.model_validate(
            {
                "version": 1,
                "tables": {
                    "reward": {
                        "primary_key": ["id"],
                        "fields": {"id": {"type": "str"}, "type_1": {"type": "str"}},
                    }
                },
            }
        )
        rows = [{"__row": index, "商品名": f"商品{index}", "奖励类型": 7, "内容ID": index} for index in range(1, 5201)]
        workbook = WorkbookIR(
            source_id="feishu-value-table",
            source_type=SourceKind.FEISHU,
            sheets=[
                SheetIR(
                    name="价值表",
                    max_row=5200,
                    max_column=4,
                    headers=["商品名", "奖励类型", "内容ID"],
                    header_row=1,
                    sample_rows=rows,
                )
            ],
        )

        context = build_minimal_context(manifest, schema, [workbook], [], None)
        sheet = context["workbooks"][0]["sheets"][0]
        self.assertEqual(len(sheet["sample_rows"]), 5000)
        self.assertEqual(sheet["sample_rows_omitted"], 200)

    def test_target_table_profiles_expose_next_active_shop_values(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            config_path = tmp / "active_shop.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "active_shop"
            sheet.append(["id", "商品组", "商品"])
            sheet.append([101, 220, 9001])
            sheet.append([102, 221, 9002])
            workbook.save(config_path)
            manifest = Manifest.model_validate(
                {
                    "project": "profile-sample",
                    "mode": "supervised_write",
                    "schema_path": str(ROOT / "config" / "example.schema.json"),
                    "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx"}],
                    "config_tables": {"active_shop": {"path": str(config_path), "sheet": "active_shop"}},
                }
            )
            schema = SchemaBundle.model_validate(
                {
                    "version": 1,
                    "tables": {
                        "active_shop": {
                            "primary_key": ["id"],
                            "fields": {"id": {"type": "int"}, "商品组": {"type": "int"}, "商品": {"type": "int"}},
                        }
                    },
                }
            )

            profiles = build_target_table_profiles(manifest, schema, tmp)
            self.assertEqual(profiles["active_shop"]["next_values"]["id"], 103)
            self.assertEqual(profiles["active_shop"]["next_values"]["商品组"], 222)

    def test_target_table_profiles_use_bottom_row_for_incremental_ids(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            config_path = tmp / "active_shop.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "active_shop"
            sheet.append(["id", "商品组", "商品"])
            sheet.append([1, 10, 9001])
            sheet.append([999, 888, 9002])
            sheet.append([10, 11, 9003])
            workbook.save(config_path)
            manifest = Manifest.model_validate(
                {
                    "project": "profile-bottom-row",
                    "mode": "supervised_write",
                    "schema_path": str(ROOT / "config" / "example.schema.json"),
                    "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx"}],
                    "config_tables": {"active_shop": {"path": str(config_path), "sheet": "active_shop"}},
                }
            )
            schema = SchemaBundle.model_validate(
                {
                    "version": 1,
                    "tables": {
                        "active_shop": {
                            "primary_key": ["id"],
                            "fields": {"id": {"type": "int"}, "商品组": {"type": "int"}, "商品": {"type": "int"}},
                        }
                    },
                }
            )

            profile = build_target_table_profiles(manifest, schema, tmp)["active_shop"]
            self.assertEqual(profile["fields"]["id"]["max_next_value"], 1000)
            self.assertEqual(profile["fields"]["id"]["bottom_last_numeric"], 10)
            self.assertEqual(profile["fields"]["id"]["next_value_basis"], "bottom_last_numeric")
            self.assertEqual(profile["next_values"]["id"], 11)
            self.assertEqual(profile["next_values"]["商品组"], 12)

    def test_activity_profile_uses_last_regular_id_before_special_sections(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            config_path = tmp / "activity.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "activity"
            sheet.append(["id", "活动备注", "活动标题"])
            sheet.append([5483, "普通活动", "a"])
            sheet.append([5484, "普通活动", "b"])
            sheet.append([None, "以下为赛季活动（7位数id）", None])
            sheet.append([2985, "S3清除危险（废弃）", "season old"])
            sheet.append([20000101, "K2王国主活动", "kvk"])
            sheet.append([20001003, "S5-贸易站活动", "season"])
            workbook.save(config_path)
            manifest = Manifest.model_validate(
                {
                    "project": "profile-activity",
                    "mode": "supervised_write",
                    "schema_path": str(ROOT / "config" / "example.schema.json"),
                    "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx"}],
                    "config_tables": {"activity": {"path": str(config_path), "sheet": "activity"}},
                }
            )
            schema = SchemaBundle.model_validate(
                {
                    "version": 1,
                    "tables": {
                        "activity": {
                            "primary_key": ["id"],
                            "fields": {"id": {"type": "int"}, "活动备注": {"type": "str"}, "活动标题": {"type": "str"}},
                        }
                    },
                }
            )

            profile = build_target_table_profiles(manifest, schema, tmp)["activity"]
            self.assertEqual(profile["fields"]["id"]["activity_regular_last_numeric"], 5484)
            self.assertEqual(profile["fields"]["id"]["bottom_last_numeric"], 20001003)
            self.assertEqual(profile["fields"]["id"]["next_value_basis"], "activity_regular_section")
            self.assertEqual(profile["next_values"]["id"], 5485)

    def test_active_shop_placeholders_are_filled_from_profiles(self) -> None:
        patch = Patch.model_validate(
            {
                "patch_id": "fill-active-shop",
                "project": "profile-sample",
                "operations": [
                    {
                        "op": "insert",
                        "target_table": "activity",
                        "rows": [{"id": 1, "活动形式模块": "7|<NEW_ACTIVE_SHOP_GROUP_PAID>|605092003|<NEW_ACTIVE_SHOP_GROUP_FREE>|605092004"}],
                        "source_ref": {"workbook": "plan"},
                        "reason": "引用新商店组",
                        "confidence": 0.8,
                        "risk_level": "medium",
                    },
                    {
                        "op": "insert",
                        "target_table": "active_shop",
                        "rows": [
                            {"id": "<NEW_ACTIVE_SHOP_ID_001>", "商品组": "<NEW_ACTIVE_SHOP_GROUP_PAID>", "商品": "r1"},
                            {"id": "<NEW_ACTIVE_SHOP_ID_002>", "商品组": "<NEW_ACTIVE_SHOP_GROUP_FREE>", "商品": "r2"},
                        ],
                        "source_ref": {"workbook": "plan"},
                        "reason": "新增商店",
                        "confidence": 0.8,
                        "risk_level": "medium",
                    },
                ],
            }
        )
        context = {"target_table_profiles": {"active_shop": {"next_values": {"id": 103, "商品组": 222}}}}

        result = fill_active_shop_incremental_ids(patch, context)
        self.assertEqual(patch.operations[1].rows[0]["id"], 103)
        self.assertEqual(patch.operations[1].rows[1]["id"], 104)
        self.assertEqual(patch.operations[1].rows[0]["商品组"], 222)
        self.assertEqual(patch.operations[1].rows[1]["商品组"], 223)
        self.assertIn("7|222|605092003|223|605092004", patch.operations[0].rows[0]["活动形式模块"])
        self.assertEqual(result["filled"]["<NEW_ACTIVE_SHOP_ID_001>"], 103)

    def test_target_table_profiles_include_generic_context_and_safe_allocations(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            config_path = tmp / "reward.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "reward"
            sheet.append(["id", "group_id", "goods_id", "type", "status"])
            sheet.append([5001, 900, 30001, 1, "open"])
            sheet.append([5002, 901, 30002, 1, "open"])
            workbook.save(config_path)
            manifest = Manifest.model_validate(
                {
                    "project": "profile-generic",
                    "mode": "supervised_write",
                    "schema_path": str(ROOT / "config" / "example.schema.json"),
                    "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx"}],
                    "config_tables": {"reward": {"path": str(config_path), "sheet": "reward"}},
                }
            )
            schema = SchemaBundle.model_validate(
                {
                    "version": 1,
                    "tables": {
                        "reward": {
                            "primary_key": ["id"],
                            "group_key": "group_id",
                            "fields": {
                                "id": {"type": "int"},
                                "group_id": {"type": "int"},
                                "goods_id": {"type": "int"},
                                "type": {"type": "int"},
                                "status": {"type": "str"},
                            },
                        }
                    },
                }
            )
            dictionary = [{"target_table": "reward", "target_field": "goods_id", "id_strategy": "lookup", "reference_table": "goods", "writable": False}]

            profiles = build_target_table_profiles(manifest, schema, tmp, dictionary)
            profile = profiles["reward"]
            self.assertEqual(profile["next_values"], {"id": 5003, "group_id": 902})
            self.assertEqual(profile["fields"]["goods_id"]["reference_table"], "goods")
            self.assertIn("lookup_ref", profile["fields"]["goods_id"]["roles"])
            self.assertIn("open", profile["fields"]["status"]["enum_values"])

    def test_generic_placeholder_allocator_skips_lookup_fields(self) -> None:
        patch = Patch.model_validate(
            {
                "patch_id": "fill-generic",
                "project": "profile-generic",
                "operations": [
                    {
                        "op": "insert",
                        "target_table": "reward",
                        "rows": [{"id": "<NEW_REWARD_ID_001>", "group_id": "<NEW_REWARD_GROUP_001>", "goods_id": "<NEW_GOODS_ID_001>"}],
                        "source_ref": {"workbook": "plan"},
                        "reason": "新增奖励",
                        "confidence": 0.8,
                        "risk_level": "medium",
                    },
                    {
                        "op": "insert",
                        "target_table": "activity",
                        "rows": [{"id": 1, "reward_ref": "<NEW_REWARD_ID_001>|<NEW_REWARD_GROUP_001>|<NEW_GOODS_ID_001>"}],
                        "source_ref": {"workbook": "plan"},
                        "reason": "引用奖励",
                        "confidence": 0.8,
                        "risk_level": "medium",
                    },
                ],
            }
        )
        context = {
            "target_table_profiles": {
                "reward": {
                    "next_values": {"id": 5003, "group_id": 902},
                    "generation_summary": {"allocatable_fields": ["id", "group_id"]},
                }
            }
        }

        result = fill_incremental_placeholders(patch, context)
        self.assertEqual(patch.operations[0].rows[0]["id"], 5003)
        self.assertEqual(patch.operations[0].rows[0]["group_id"], 902)
        self.assertEqual(patch.operations[0].rows[0]["goods_id"], "<NEW_GOODS_ID_001>")
        self.assertEqual(patch.operations[1].rows[0]["reward_ref"], "5003|902|<NEW_GOODS_ID_001>")
        self.assertIn("<NEW_REWARD_ID_001>", result["filled"])
        self.assertEqual(result["skipped_fields"][0]["field"], "goods_id")

    def test_allocator_corrects_concrete_activity_and_active_shop_ids_from_bottom_profiles(self) -> None:
        patch = Patch.model_validate(
            {
                "patch_id": "correct-concrete-ids",
                "project": "profile-bottom-row",
                "operations": [
                    {
                        "op": "insert",
                        "target_table": "activity",
                        "rows": [{"id": "2441002", "活动形式模块": "7|10004|605092003|7|10005|605092004"}],
                        "source_ref": {"workbook": "plan"},
                        "reason": "AI used max id",
                        "confidence": 0.8,
                        "risk_level": "medium",
                    },
                    {
                        "op": "insert",
                        "target_table": "active_shop",
                        "rows": [
                            {"id": "510133010", "商品组": "10004", "商品": "605300001"},
                            {"id": "510133011", "商品组": "10004", "商品": "605300002"},
                            {"id": "510133012", "商品组": "10005", "商品": "605300003"},
                        ],
                        "source_ref": {"workbook": "plan"},
                        "reason": "AI used max id",
                        "confidence": 0.8,
                        "risk_level": "medium",
                    },
                    {
                        "op": "insert",
                        "target_table": "reward",
                        "rows": [{"id": "605300001"}],
                        "source_ref": {"workbook": "plan"},
                        "reason": "reward uses a domain date rule",
                        "confidence": 0.8,
                        "risk_level": "medium",
                    },
                ],
            }
        )
        context = {
            "target_table_profiles": {
                "activity": {
                    "next_values": {"id": 5466},
                    "generation_summary": {"allocatable_fields": ["id"]},
                    "fields": {"id": {"allocation_role": "primary_key", "id_strategy": "new", "next_value_basis": "activity_regular_section"}},
                },
                "active_shop": {
                    "next_values": {"id": 5005, "商品组": 532},
                    "generation_summary": {"allocatable_fields": ["id", "商品组"]},
                    "fields": {
                        "id": {"allocation_role": "primary_key", "id_strategy": "", "next_value_basis": "bottom_last_numeric"},
                        "商品组": {"allocation_role": "group_key", "id_strategy": "", "next_value_basis": "bottom_last_numeric"},
                    },
                },
                "reward": {
                    "next_values": {"id": 1012303},
                    "generation_summary": {"allocatable_fields": ["id"]},
                    "fields": {"id": {"allocation_role": "primary_key", "id_strategy": "new_or_reuse", "next_value_basis": "bottom_last_numeric"}},
                },
            }
        }

        result = fill_incremental_placeholders(patch, context)
        self.assertEqual(patch.operations[0].rows[0]["id"], 5466)
        self.assertEqual(patch.operations[0].rows[0]["活动形式模块"], "7|532|605092003|7|533|605092004")
        active_rows = patch.operations[1].rows
        self.assertEqual([row["id"] for row in active_rows], [5005, 5006, 5007])
        self.assertEqual([row["商品组"] for row in active_rows], [532, 532, 533])
        self.assertEqual(patch.operations[2].rows[0]["id"], "605300001")
        self.assertGreaterEqual(len(result["corrected_fields"]), 5)

    def test_structured_correction_is_reused_in_experience_context(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            manifest = Manifest.model_validate(
                {
                    "project": "correction-sample",
                    "mode": "supervised_write",
                    "schema_path": str(ROOT / "config" / "example.schema.json"),
                    "run_instruction": "兑换商店活动，价格字段需要复核。",
                    "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx"}],
                    "target_tables": ["exchange"],
                }
            )
            schema = SchemaBundle.model_validate(
                {
                    "version": 1,
                    "tables": {
                        "exchange": {
                            "primary_key": ["id"],
                            "fields": {"id": {"type": "int"}, "支付价格": {"type": "int"}},
                        }
                    },
                }
            )
            patch_obj = Patch.model_validate(
                {
                    "patch_id": "patch_correction",
                    "project": "correction-sample",
                    "operations": [
                        {
                            "op": "insert",
                            "target_table": "exchange",
                            "rows": [{"id": 1, "支付价格": 100}],
                            "reason": "unit",
                            "confidence": 0.8,
                        }
                    ],
                }
            )
            record = {"target_tables": ["exchange"], "validation_summary": {}}
            review = local_case_review("价格字段不能直接取原价，要取规划里的现价。", record)
            correction = build_structured_correction(manifest, patch_obj, "价格字段不能直接取原价，要取规划里的现价。", review, record)
            save_structured_correction(tmp, correction)
            workbook = WorkbookIR(
                source_id="plan",
                source_type=SourceKind.LOCAL_EXCEL,
                sheets=[SheetIR(name="兑换规划", max_row=2, max_column=2, headers=["价格", "商品"], header_row=1, sample_rows=[{"价格": 100}])],
            )
            experience = build_experience_context(tmp, manifest, schema, [workbook], {})
            self.assertTrue(experience["structured_corrections"])
            self.assertIn("现价", experience["structured_corrections"][0]["correct_practice"])

    def test_experience_summary_ai_uses_review_json_shape(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "summary-ai-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx", "role": "planning"}],
                "ai": {"provider": "deepseek_v4_pro"},
            }
        )
        response_payload = {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "summary_title": "兑换商店经验",
                                "review_text": "商品名 -> goods.name",
                                "activity_templates": [],
                                "field_mappings": [],
                                "personal_rules": [],
                                "questions": [],
                                "risk_notes": [],
                                "conflicts": [],
                            }
                        )
                    }
                }
            ]
        }
        captured = {}

        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                return json.dumps(response_payload).encode("utf-8")

        def fake_urlopen(request, timeout):
            captured["url"] = request.full_url
            captured["body"] = json.loads(request.data.decode("utf-8"))
            return FakeResponse()

        with patch.dict(os.environ, {"BASEAI_API_KEY": "unit-key"}, clear=True):
            with patch("urllib.request.urlopen", side_effect=fake_urlopen):
                summary = call_experience_summary_ai(manifest, {"raw_experience": "兑换商店经验"})

        self.assertEqual(summary["summary_title"], "兑换商店经验")
        self.assertEqual(captured["url"], "https://baseai.rivergame.net/v1/chat/completions")
        self.assertEqual(captured["body"]["response_format"], {"type": "json_object"})
        self.assertIn("review_text", captured["body"]["messages"][0]["content"])
        self.assertIn("conflicts", captured["body"]["messages"][0]["content"])

    def test_activity_template_matching_builds_config_plan(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            manifest = Manifest.model_validate(
                {
                    "project": "plan-sample",
                    "mode": "supervised_write",
                    "schema_path": "schema.json",
                    "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "planning.xlsx", "role": "planning"}],
                    "target_tables": ["activity"],
                }
            )
            schema = SchemaBundle.model_validate(
                {
                    "tables": {
                        "activity": {"primary_key": ["id"], "fields": {"id": {}, "title": {}, "form_list": {}}},
                        "active_shop": {"primary_key": ["id"], "fields": {"id": {}, "group": {}, "goods": {}}},
                        "exchange": {"primary_key": ["id"], "fields": {"id": {}, "activity_id": {}, "price": {}}},
                        "reward": {"primary_key": ["id"], "fields": {"id": {}}},
                        "goods": {"primary_key": ["item_id"], "fields": {"item_id": {}, "name": {}}},
                        "key": {"primary_key": ["key"], "fields": {"key": {}, "text": {}}},
                    }
                }
            )
            workbook = WorkbookIR(
                source_id="plan",
                source_type=SourceKind.LOCAL_EXCEL,
                sheets=[
                    SheetIR(
                        name="exchange shop planning",
                        max_row=3,
                        max_column=3,
                        header_row=1,
                        headers=["item_id", "price", "reward"],
                        sample_rows=[{"item_id": 3001, "price": 68, "reward": "3001*10"}],
                    )
                ],
            )
            experience = build_experience_context(
                tmp,
                manifest,
                schema,
                [workbook],
                {"recommended_tables": ["active_shop", "reward"], "relations": []},
            )
            plan = experience["config_plan"]
            self.assertIn("exchange", plan["activity_template_id"])
            self.assertIn("active_shop", plan["all_recommended_tables"])
            self.assertIn("reward", plan["all_recommended_tables"])
            self.assertTrue(plan["matched_field_mappings"])
            self.assertTrue(plan["pending_confirmations"])

    def test_patch_learning_also_records_case_example(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "case-sample",
                "mode": "supervised_write",
                "schema_path": "schema.json",
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "planning.xlsx", "role": "planning"}],
            }
        )
        patch_obj = Patch(
            patch_id="patch_case",
            project="case-sample",
            operations=[
                {
                    "op": "update",
                    "target_table": "activity",
                    "match": {"id": 1},
                    "set": {"title": "demo"},
                    "reason": "unit test",
                    "confidence": 0.8,
                    "risk_level": "low",
                }
            ],
        )
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            case = append_case_from_patch(tmp, manifest, patch_obj, "accepted", "kept title style")
            cases = [json.loads(line) for line in (tmp / ".knowledge" / "case_examples.jsonl").read_text(encoding="utf-8").splitlines()]
        self.assertEqual(case["target_tables"], ["activity"])
        self.assertEqual(cases[0]["decision"], "accepted")

    def test_config_plan_auto_expands_generation_target_tables(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "auto-target-sample",
                "schema_path": "schema.json",
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "planning.xlsx", "role": "planning"}],
                "target_tables": ["activity", "active_shop"],
            }
        )
        schema = SchemaBundle.model_validate(
            {
                "tables": {
                    "activity": {"primary_key": ["id"], "fields": {"id": {}, "form_list": {}}},
                    "active_shop": {"primary_key": ["id"], "fields": {"id": {}, "group": {}}},
                    "exchange": {"primary_key": ["id"], "fields": {"id": {}, "activity_id": {}, "price": {}}},
                    "reward": {"primary_key": ["id"], "fields": {"id": {}}},
                    "goods": {"primary_key": ["item_id"], "fields": {"item_id": {}, "name": {}}},
                    "key": {"primary_key": ["key"], "fields": {"key": {}, "text": {}}},
                }
            }
        )
        plan = {
            "relation_chain": ["activity", "active_shop", "exchange", "reward", "goods", "key"],
            "required_fields": {"exchange": ["id"], "goods": ["item_id"]},
            "recommended_target_tables": ["exchange", "goods", "missing_table"],
        }

        tables, included = _auto_expand_generation_tables(manifest, schema, plan)

        self.assertEqual(tables, ["activity", "active_shop", "exchange", "reward", "goods", "key"])
        self.assertEqual(included, ["exchange", "reward", "goods", "key"])

    def test_config_root_discovers_tables(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            planning = tmp / "planning.xlsx"
            config_dir = tmp / "configs"
            config_dir.mkdir()
            config = config_dir / "tables.xlsx"
            misleading = config_dir / "shop_pack_config.xlsx"
            make_planning(planning)
            make_config(config)
            misleading_workbook = Workbook()
            misleading_sheet = misleading_workbook.active
            misleading_sheet.title = "说明"
            misleading_sheet["A1"] = "这个文件名像配置表，但 sheet 不是数据表"
            misleading_workbook.save(misleading)
            manifest = {
                "project": "root-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "run_root": str(tmp / ".runs"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": str(planning), "role": "planning"}],
                "config_roots": [{"path": str(config_dir), "recursive": True}],
                "habit_store": str(tmp / ".knowledge" / "habits.jsonl"),
            }
            manifest_path = tmp / "manifest-root.json"
            write_json(manifest_path, manifest)

            manifest_model, schema, _, context = analyze_manifest(manifest_path, tmp, "analysis")
            self.assertIn("shop_pack_config", manifest_model.config_tables)
            self.assertIn("reward_item_config", manifest_model.config_tables)
            self.assertEqual(context["config_discovery"]["matched"]["shop_pack_config"]["source"], "sheet_name")
            self.assertTrue(any(item["name"] == "说明" for item in context["config_discovery"]["skipped_sheets"]))
            patch = make_stub_patch(manifest_model, schema, context, str(tmp))
            self.assertEqual([op.op for op in patch.operations], ["update", "insert"])

    def test_config_root_does_not_match_by_file_name(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            planning = tmp / "planning.xlsx"
            config_dir = tmp / "configs"
            config_dir.mkdir()
            make_planning(planning)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["pack_id", "name"])
            sheet.append([1001, "文件名不能代表表名"])
            workbook.save(config_dir / "shop_pack_config.xlsx")
            manifest = {
                "project": "file-name-is-not-table",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "run_root": str(tmp / ".runs"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": str(planning), "role": "planning"}],
                "config_roots": [{"path": str(config_dir), "recursive": True}],
                "habit_store": str(tmp / ".knowledge" / "habits.jsonl"),
            }
            manifest_path = tmp / "manifest-root.json"
            write_json(manifest_path, manifest)

            manifest_model, _, _, context = analyze_manifest(manifest_path, tmp, "analysis")
            self.assertNotIn("shop_pack_config", manifest_model.config_tables)
            self.assertIn("shop_pack_config", context["config_discovery"]["unmatched_tables"])

    def test_schema_scan_builds_schema_draft_from_sheet_names(self) -> None:
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            config_dir = tmp / "configs"
            config_dir.mkdir()
            config = config_dir / "multi.xlsx"
            make_config(config)
            workbook = load_workbook(config)
            workbook["shop_pack_config"]["F2"] = datetime(2026, 6, 1, 5, 0, 0)
            note = workbook.create_sheet("说明")
            note["A1"] = "非数据页"
            workbook.save(config)
            manifest = Manifest.model_validate(
                {
                    "project": "schema-scan-sample",
                    "mode": "supervised_write",
                    "schema_path": str(ROOT / "config" / "example.schema.json"),
                    "run_root": str(tmp / ".runs"),
                    "planning_sources": [{"id": "dummy", "kind": "local_excel", "path": str(config), "role": "planning"}],
                    "config_roots": [{"path": str(config_dir), "recursive": True}],
                    "habit_store": str(tmp / ".knowledge" / "habits.jsonl"),
                }
            )
            run_dir = tmp / ".runs" / "schema-scan"
            run_dir.mkdir(parents=True)
            result = scan_config_schema(manifest, tmp, run_dir)
            schema_draft = result["schema_draft"]
            report = result["report"]
            self.assertIn("shop_pack_config", schema_draft["tables"])
            self.assertIn("reward_item_config", schema_draft["tables"])
            self.assertIn("pack_id", schema_draft["tables"]["shop_pack_config"]["fields"])
            self.assertTrue(any(item["name"] == "说明" for item in report["skipped_sheets"]))
            self.assertTrue((run_dir / "schema-draft.json").exists())
            saved_report = read_json(run_dir / "schema-scan.json")
            self.assertEqual(saved_report["tables"]["shop_pack_config"]["sample_rows"][0]["start_time"], "2026-06-01T05:00:00")

    def test_relation_scan_discovers_id_group_reward_and_key_links(self) -> None:
        self.assertEqual(split_reference_values("[100, 101]"), ["100", "101"])
        self.assertEqual(split_reference_values("100|101,102"), ["100", "101", "102"])
        with tempfile.TemporaryDirectory() as raw:
            tmp = Path(raw)
            config_dir = tmp / "configs"
            config_dir.mkdir()
            config = config_dir / "relations.xlsx"
            make_relation_config(config)
            manifest = Manifest.model_validate(
                {
                    "project": "relation-sample",
                    "mode": "supervised_write",
                    "schema_path": str(tmp / ".runs" / "schema-scan" / "schema-draft.json"),
                    "run_root": str(tmp / ".runs"),
                    "planning_sources": [{"id": "dummy", "kind": "local_excel", "path": str(config), "role": "planning"}],
                    "config_roots": [{"path": str(config_dir), "recursive": True}],
                    "target_tables": ["activity"],
                    "habit_store": str(tmp / ".knowledge" / "habits.jsonl"),
                }
            )
            scan_dir = tmp / ".runs" / "schema-scan"
            scan_dir.mkdir(parents=True)
            scan_config_schema(manifest, tmp, scan_dir)
            schema = load_schema(scan_dir / "schema-draft.json")

            relation_dir = tmp / ".runs" / "relation-scan"
            relation_dir.mkdir(parents=True)
            result = scan_relationships(manifest, schema, tmp, relation_dir, scan_dir / "schema-scan.json")
            relation_keys = {
                (item["from_table"], item["from_field"], item["to_table"], item["to_field"])
                for item in result["relations"]
            }

            self.assertIn(("activity", "form_list", "active_shop", "group"), relation_keys)
            self.assertIn(("active_shop", "goods", "reward", "id"), relation_keys)
            self.assertIn(("active_shop", "title_key", "key", "key"), relation_keys)
            self.assertIn("active_shop", result["recommended_tables"])
            self.assertIn("reward", result["recommended_tables"])
            self.assertTrue((relation_dir / "relationship-map.json").exists())

    def test_feishu_sheet_source_becomes_workbook_ir(self) -> None:
        source = PlanningSource(id="plan", kind="feishu", url="https://rivergame.feishu.cn/wiki/demo?sheet=abc", role="planning")
        payload = FeishuSourcePayload(kind="sheet", title="飞书规划表", values=[["礼包ID", "礼包名称"], [1001, "每日礼包"]])
        with tempfile.TemporaryDirectory() as raw, patch("ai_meta_agent.workbook_ir.read_feishu_source", return_value=payload) as reader:
            workbook = load_source_ir(source, Path(raw))
        reader.assert_called_once()
        self.assertEqual(workbook.source_id, "plan")
        self.assertEqual(workbook.sheets[0].name, "飞书规划表")
        self.assertEqual(workbook.sheets[0].header_row, 1)
        self.assertEqual(workbook.sheets[0].sample_rows[0]["礼包ID"], 1001)

    def test_feishu_doc_source_becomes_text_sheet(self) -> None:
        source = PlanningSource(id="doc", kind="feishu", url="https://rivergame.feishu.cn/docx/demo", role="planning")
        payload = FeishuSourcePayload(kind="doc", title="活动规划", text="活动时间：6月1日\n奖励：金币")
        with tempfile.TemporaryDirectory() as raw, patch("ai_meta_agent.workbook_ir.read_feishu_source", return_value=payload):
            workbook = load_source_ir(source, Path(raw))
        sheet = workbook.sheets[0]
        self.assertEqual(sheet.name, "活动规划")
        self.assertEqual(sheet.headers[:2], ["section", "content"])
        self.assertIn("活动时间", sheet.sample_rows[0]["content"])

    def test_feishu_sheet_read_falls_back_to_chunks_when_payload_is_too_large(self) -> None:
        calls: list[str] = []

        def fake_run_lark_cli(_lark_cli: str, args: list[str], _label: str, _cwd: Path) -> dict[str, Any]:
            if args[:3] == ["wiki", "spaces", "get_node"]:
                return {"data": {"node": {"obj_type": "sheet", "obj_token": "sheet-token", "title": "value-table"}}}
            range_name = args[args.index("--range") + 1]
            calls.append(range_name)
            if range_name == "A1:ZZ1200":
                raise RuntimeError("API call failed: [90221] data exceeded 10485760 bytes.")
            values_by_range = {
                "A1:ZZ500": [["item_name", "reward_type", "content_id"], ["ruby", 7, 323]],
                "A501:ZZ1000": [["gold", 8, 999]],
                "A1001:ZZ1200": [],
            }
            return {"data": {"valueRange": {"range": range_name, "values": values_by_range[range_name]}}}

        with tempfile.TemporaryDirectory() as raw:
            with patch("ai_meta_agent.feishu.resolve_lark_cli", return_value="lark-cli.exe"):
                with patch("ai_meta_agent.feishu.run_lark_cli", side_effect=fake_run_lark_cli):
                    payload = read_feishu_sheet(
                        "https://rivergame.feishu.cn/wiki/demo?sheet=abc",
                        Path(raw),
                        range_name="A1:ZZ1200",
                    )

        self.assertEqual(calls, ["A1:ZZ1200", "A1:ZZ500", "A501:ZZ1000", "A1001:ZZ1200"])
        self.assertEqual(payload.values[0], ["item_name", "reward_type", "content_id"])
        self.assertIn(["gold", 8, 999], payload.values)
        self.assertTrue(any("分块读取" in notice for notice in payload.notices))

    def test_planning_items_resolve_reward_type_and_id_from_value_table(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "item-resolution-sample",
                "schema_path": "schema.json",
                "planning_sources": [
                    {"id": "plan", "kind": "feishu", "url": "https://rivergame.feishu.cn/wiki/plan?sheet=a", "role": "planning"},
                    {"id": "value", "kind": "feishu", "url": "https://rivergame.feishu.cn/wiki/value?sheet=b", "role": "item_base"},
                ],
            }
        )
        workbooks = [
            WorkbookIR(
                source_id="plan",
                source_type=SourceKind.FEISHU,
                sheets=[
                    SheetIR(
                        name="规划表",
                        max_row=2,
                        max_column=2,
                        headers=["商品名称", "价格"],
                        header_row=1,
                        sample_rows=[{"__row": 2, "商品名称": "克拉肯", "价格": 100}],
                    )
                ],
            ),
            WorkbookIR(
                source_id="value",
                source_type=SourceKind.FEISHU,
                sheets=[
                    SheetIR(
                        name="基础价值表",
                        max_row=2,
                        max_column=4,
                        headers=["商品名称", "奖励类型", "内容ID", "数量"],
                        header_row=1,
                        sample_rows=[{"__row": 2, "商品名称": "克拉肯", "奖励类型": 7, "内容ID": 323, "数量": 1}],
                    )
                ],
            ),
        ]
        result = resolve_planning_items(manifest, workbooks)
        self.assertTrue(result["enabled"])
        self.assertEqual(result["summary"]["matched"], 1)
        self.assertEqual(result["matches"][0]["reward_type"], 7)
        self.assertEqual(result["matches"][0]["content_id"], 323)

    def test_planning_items_infer_weak_value_table_columns(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "weak-value-table",
                "schema_path": "schema.json",
                "planning_sources": [
                    {"id": "plan", "kind": "feishu", "url": "https://rivergame.feishu.cn/wiki/plan?sheet=a", "role": "planning"},
                    {"id": "value", "kind": "feishu", "url": "https://rivergame.feishu.cn/wiki/value?sheet=b", "role": "item_base"},
                ],
            }
        )
        workbooks = [
            WorkbookIR(
                source_id="plan",
                source_type=SourceKind.FEISHU,
                sheets=[
                    SheetIR(
                        name="规划表",
                        max_row=2,
                        max_column=2,
                        headers=["商品名称", "价格"],
                        header_row=1,
                        sample_rows=[{"__row": 2, "商品名称": "克拉肯", "价格": 100}],
                    )
                ],
            ),
            WorkbookIR(
                source_id="value",
                source_type=SourceKind.FEISHU,
                sheets=[
                    SheetIR(
                        name="配置转换表",
                        max_row=4,
                        max_column=4,
                        headers=["部族灵魂石", "33", "2", "142"],
                        header_row=6,
                        sample_rows=[
                            {"__row": 7, "部族灵魂石": "克拉肯", "33": 33, "2": 7, "142": 323},
                            {"__row": 8, "部族灵魂石": "自然灵魂石", "33": 33, "2": 2, "142": 143},
                            {"__row": 9, "部族灵魂石": "武器进阶石", "33": 50, "2": 2, "142": 144},
                        ],
                    )
                ],
            ),
        ]

        result = resolve_planning_items(manifest, workbooks)

        self.assertEqual(result["summary"]["matched"], 1)
        self.assertEqual(result["matches"][0]["reward_type"], 7)
        self.assertEqual(result["matches"][0]["content_id"], 323)
        self.assertEqual(result["summary"]["inferred_item_base_sheets"], 1)

    def test_ai_context_omits_value_rows_and_reports_budget(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "optimized-context",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [
                    {"id": "plan", "kind": "local_excel", "path": "dummy.xlsx", "role": "planning"},
                    {"id": "value", "kind": "local_excel", "path": "value.xlsx", "role": "item_base"},
                ],
            }
        )
        schema = SchemaBundle.model_validate(
            {
                "version": 1,
                "tables": {
                    "reward": {
                        "primary_key": ["id"],
                        "fields": {"id": {"type": "str"}, "type_1": {"type": "str"}, "reward_1": {"type": "str"}},
                    }
                },
            }
        )
        workbooks = [
            WorkbookIR(
                source_id="plan",
                source_type=SourceKind.LOCAL_EXCEL,
                sheets=[
                    SheetIR(
                        name="规划表",
                        max_row=2,
                        max_column=2,
                        headers=["商品名称", "价格"],
                        header_row=1,
                        sample_rows=[{"__row": 2, "商品名称": "商品42", "价格": 100}],
                    )
                ],
            ),
            WorkbookIR(
                source_id="feishu-value-table",
                source_type=SourceKind.FEISHU,
                sheets=[
                    SheetIR(
                        name="价值表",
                        max_row=5200,
                        max_column=4,
                        headers=["商品名", "奖励类型", "内容ID"],
                        header_row=1,
                        sample_rows=[{"__row": index, "商品名": f"商品{index}", "奖励类型": 7, "内容ID": index} for index in range(1, 5201)],
                    )
                ],
            ),
        ]
        resolution = resolve_planning_items(manifest, workbooks)
        context = build_minimal_context(manifest, schema, workbooks, [], None, resolution)

        optimized = optimize_context_for_ai(context)
        budget = build_context_budget(context, optimized)
        value_sheet = [workbook for workbook in optimized["workbooks"] if workbook["source_id"] == "feishu-value-table"][0]["sheets"][0]

        self.assertEqual(value_sheet["sample_rows"], [])
        self.assertEqual(budget["rows"]["value_sample_rows_before"], 5000)
        self.assertEqual(budget["rows"]["value_sample_rows_sent_to_ai"], 0)
        self.assertLess(budget["optimized"]["bytes"], budget["original"]["bytes"])
        self.assertIn("resolved_items", optimized)

    def test_hard_fast_budget_limits_model_tables_and_knowledge(self) -> None:
        tables = {}
        profiles = {}
        target_tables = ["activity", "active_shop", "reward", "goods", "key", "jump", "exchange", "activity_drop"]
        for table_name in target_tables:
            field_names = ["id", "group", "reward", "goods", "price", *[f"field_{index}" for index in range(40)]]
            tables[table_name] = {
                "primary_key": ["id"],
                "group_key": "group",
                "field_names": field_names,
                "required_fields": field_names[:12],
                "field_types": {field: "int" for field in field_names[:20]},
            }
            profiles[table_name] = {
                "primary_key": ["id"],
                "group_key": "group",
                "next_values": {"id": 1000, "group": 80},
                "fields": {
                    field: {"field": field, "next_value": 1000, "sample_values": ["x" * 80], "enum_values": [1, 2, 3, 4]}
                    for field in field_names
                },
            }
        context = {
            "target_tables": target_tables,
            "schema": {"tables": tables},
            "target_table_profiles": profiles,
            "planning_evidence": [
                {
                    "source_id": "plan",
                    "sheet": "planning",
                    "rows": [{"__row": index, "商品": f"item-{index}", "价格": index, "长说明": "x" * 220} for index in range(80)],
                }
            ],
            "resolved_items": [
                {
                    "product_name": f"item-{index}",
                    "reward_type": 7,
                    "content_id": index,
                    "planning_ref": {"workbook": "plan", "sheet": "planning", "row": index, "url": "https://example.com/" + "x" * 120},
                }
                for index in range(24)
            ],
            "config_plan": {
                "activity_type": "兑换商店活动",
                "relation_chain": ["activity", "active_shop", "reward", "goods", "key", "jump"],
                "id_strategy": {"template_rule": "active_shop 和 activity 都按最新 ID 递增。" * 80},
                "required_fields": {"activity": ["id", "type"], "active_shop": ["id", "goods"]},
            },
            "similar_cases": [{"case_id": str(index), "correction": "x" * 1200} for index in range(6)],
            "matched_rules": [{"title": "rule", "text": "x" * 1200, "applies_to_tables": target_tables} for _ in range(6)],
            "structured_corrections": [{"correct_practice": "x" * 1200, "target_tables": target_tables} for _ in range(4)],
        }

        compact = enforce_fast_context_budget(context)
        budget = build_context_budget(context, compact)

        self.assertIn((compact["context_optimization"] or {})["mode"], {"fast_budget_local_evidence_first", "hard_fast_budget_local_evidence_first"})
        self.assertLessEqual(budget["optimized"]["bytes"], 32 * 1024)
        self.assertEqual(compact["ai_target_tables"], ["activity", "active_shop", "reward", "goods"])
        self.assertLessEqual(budget["rows"]["planning_evidence_rows_sent_to_ai"], 22)
        self.assertLessEqual(len(compact["similar_cases"]), 1)

    def test_compact_profiles_keep_id_allocation_evidence(self) -> None:
        profiles = {
            "activity": {
                "sheet": "activity",
                "header_row": 1,
                "row_count": 6000,
                "primary_key": ["id"],
                "group_key": None,
                "generation_summary": {"allocatable_fields": ["id"]},
                "next_values": {"id": 5861},
                "fields": {
                    "id": {
                        "field": "id",
                        "allocation_role": "primary_key",
                        "next_value": 5861,
                        "next_value_basis": "activity_regular_section",
                        "activity_regular_last_numeric": 5860,
                        "activity_regular_last_numeric_row": 6411,
                    },
                    "备注": {"field": "备注", "sample_values": ["旧活动"]},
                },
                "tail_rows": [{"__row": 6411, "id": 5860, "备注": "常规活动"}, {"__row": 6420, "id": 20001003, "备注": "以下为赛季活动"}],
            }
        }

        compact = compact_target_table_profiles(profiles)

        self.assertEqual(compact["activity"]["next_values"]["id"], 5861)
        self.assertEqual(compact["activity"]["fields"]["id"]["next_value_basis"], "activity_regular_section")
        self.assertNotIn("备注", compact["activity"]["fields"])

    def test_company_bi_model_provider_choices_share_baseai_endpoint(self) -> None:
        response_payload = {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "patch_id": "patch_company_bi",
                                "project": "provider-sample",
                                "mode": "supervised_write",
                                "operations": [],
                                "generated_by": "ai-meta-agent",
                            }
                        )
                    }
                }
            ]
        }
        captured = []

        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                return json.dumps(response_payload).encode("utf-8")

        def fake_urlopen(request, timeout):
            captured.append({"url": request.full_url, "body": json.loads(request.data.decode("utf-8"))})
            return FakeResponse()

        providers = {
            "baseai": "gpt-5.5",
            "gemini": "gemini-3.1-pro-preview",
            "claude": "claude-opus-4-8",
        }
        with patch.dict(os.environ, {"BASEAI_API_KEY": "unit-key"}, clear=True):
            with patch("urllib.request.urlopen", side_effect=fake_urlopen):
                for provider in providers:
                    manifest = Manifest.model_validate(
                        {
                            "project": "provider-sample",
                            "mode": "supervised_write",
                            "schema_path": str(ROOT / "config" / "example.schema.json"),
                            "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx", "role": "planning"}],
                            "ai": {"provider": provider},
                        }
                    )
                    generated = call_baseai(manifest, {"project": "provider-sample"})
                    self.assertEqual(generated.patch_id, "patch_company_bi")

        self.assertEqual([item["body"]["model"] for item in captured], list(providers.values()))
        self.assertTrue(all(item["url"] == "https://baseai.rivergame.net/v1/chat/completions" for item in captured))
        self.assertTrue(all(item["body"]["max_tokens"] == 7000 for item in captured))
        self.assertIn("temperature", captured[0]["body"])
        self.assertIn("temperature", captured[1]["body"])
        self.assertNotIn("temperature", captured[2]["body"])

    def test_deepseek_provider_uses_company_bi_request(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "deepseek-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx", "role": "planning"}],
                "ai": {"provider": "deepseek_v4_pro"},
            }
        )
        response_payload = {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "patch_id": "patch_deepseek",
                                "project": "deepseek-sample",
                                "mode": "supervised_write",
                                "operations": [],
                                "generated_by": "ai-meta-agent",
                            }
                        )
                    }
                }
            ],
            "usage": {"prompt_tokens": 1, "completion_tokens": 1},
        }
        captured = {}

        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                return json.dumps(response_payload).encode("utf-8")

        def fake_urlopen(request, timeout):
            captured["url"] = request.full_url
            captured["timeout"] = timeout
            captured["body"] = json.loads(request.data.decode("utf-8"))
            captured["authorization"] = request.headers.get("Authorization")
            return FakeResponse()

        with tempfile.TemporaryDirectory() as raw:
            raw_response = Path(raw) / "ai-response.json"
            with patch.dict(os.environ, {"BASEAI_API_KEY": "unit-key"}, clear=True):
                with patch("urllib.request.urlopen", side_effect=fake_urlopen):
                    generated = call_baseai(manifest, {"project": "deepseek-sample"}, raw_response)
            self.assertTrue(raw_response.exists())

        self.assertEqual(generated.patch_id, "patch_deepseek")
        self.assertEqual(captured["url"], "https://baseai.rivergame.net/v1/chat/completions")
        self.assertEqual(captured["body"]["model"], "deepseek-v4-pro")
        self.assertEqual(captured["body"]["max_tokens"], 7000)
        self.assertNotIn("thinking", captured["body"])
        self.assertEqual(captured["body"]["response_format"], {"type": "json_object"})
        self.assertEqual(captured["authorization"], "Bearer unit-key")

    def test_ai_generation_budget_can_be_overridden_per_provider(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "deepseek-budget-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx", "role": "planning"}],
                "ai": {"provider": "deepseek_v4_pro"},
            }
        )
        response_payload = {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "patch_id": "patch_budget",
                                "project": "deepseek-budget-sample",
                                "mode": "supervised_write",
                                "operations": [],
                                "generated_by": "ai-meta-agent",
                            }
                        )
                    }
                }
            ]
        }
        captured = {}

        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                return json.dumps(response_payload).encode("utf-8")

        def fake_urlopen(request, timeout):
            captured["body"] = json.loads(request.data.decode("utf-8"))
            return FakeResponse()

        env = {
            "BASEAI_API_KEY": "unit-key",
            "AI_MAX_OUTPUT_TOKENS": "9000",
            "DEEPSEEK_MAX_OUTPUT_TOKENS": "4321",
            "DEEPSEEK_REASONING_EFFORT": "low",
        }
        with patch.dict(os.environ, env, clear=True):
            with patch("urllib.request.urlopen", side_effect=fake_urlopen):
                generated = call_baseai(manifest, {"project": "deepseek-budget-sample"})

        self.assertEqual(generated.patch_id, "patch_budget")
        self.assertEqual(captured["body"]["max_tokens"], 4321)
        self.assertEqual(captured["body"]["reasoning_effort"], "low")

    def test_deepseek_provider_requires_company_bi_key(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "deepseek-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx", "role": "planning"}],
                "ai": {"provider": "deepseek-v4-pro"},
            }
        )
        with patch.dict(os.environ, {}, clear=True):
            with self.assertRaisesRegex(RuntimeError, "BASEAI_API_KEY"):
                call_baseai(manifest, {})

    def test_relationship_ai_explainer_uses_compact_relation_context(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "relation-ai-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx", "role": "planning"}],
                "ai": {"provider": "deepseek_v4_pro"},
            }
        )
        response_payload = {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "summary": "activity links to active_shop",
                                "recommended_tables": ["active_shop"],
                                "relation_notes": [],
                                "needs_confirmation": [],
                            }
                        )
                    }
                }
            ]
        }
        captured = {}

        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                return json.dumps(response_payload).encode("utf-8")

        def fake_urlopen(request, timeout):
            captured["url"] = request.full_url
            captured["body"] = json.loads(request.data.decode("utf-8"))
            return FakeResponse()

        with patch.dict(os.environ, {"BASEAI_API_KEY": "unit-key"}, clear=True):
            with patch("urllib.request.urlopen", side_effect=fake_urlopen):
                review = call_relationship_ai(
                    manifest,
                    {
                        "relations": [
                            {
                                "from_table": "activity",
                                "from_field": "form_list",
                                "to_table": "active_shop",
                                "to_field": "group",
                            }
                        ]
                    },
                )
        self.assertEqual(review["recommended_tables"], ["active_shop"])
        self.assertEqual(captured["url"], "https://baseai.rivergame.net/v1/chat/completions")
        self.assertEqual(captured["body"]["response_format"], {"type": "json_object"})

    def test_empty_draft_diagnostics_explains_missing_mapping(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "diagnostic-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx", "role": "planning"}],
            }
        )
        patch_obj = Patch(patch_id="empty", project="diagnostic-sample", operations=[])
        context = {
            "target_tables": ["activity", "reward"],
            "workbooks": [
                {
                    "source_id": "plan",
                    "source_type": "local_excel",
                    "sheets": [
                        {
                            "name": "活动规划",
                            "max_row": 10,
                            "max_column": 4,
                            "header_row": 1,
                            "headers": ["道具名称", "数量", "价格"],
                            "sample_rows": [{"道具名称": "金币", "数量": 10}],
                        }
                    ],
                }
            ],
            "schema": {
                "tables": {
                    "activity": {"primary_key": ["id"], "fields": {"id": {}, "活动标题": {}, "form_list": {}}},
                    "reward": {"primary_key": ["id"], "fields": {"id": {}, "奖励": {}, "数量": {}}},
                }
            },
            "relationship_map": {
                "summary": {"relation_count": 3, "high_confidence_count": 2},
                "recommended_tables": ["goods"],
            },
        }
        diagnostics = build_draft_diagnostics(manifest, context, patch_obj)
        compact = compact_draft_diagnostic_context(manifest, context, patch_obj)

        self.assertEqual(diagnostics["status"], "empty")
        self.assertIn("goods", diagnostics["suggested_target_tables"])
        self.assertTrue(diagnostics["missing_information"])
        self.assertIn("activity", compact["schema_tables"])

    def test_draft_diagnostics_ai_uses_compact_context(self) -> None:
        manifest = Manifest.model_validate(
            {
                "project": "diagnostic-ai-sample",
                "mode": "supervised_write",
                "schema_path": str(ROOT / "config" / "example.schema.json"),
                "planning_sources": [{"id": "plan", "kind": "local_excel", "path": "dummy.xlsx", "role": "planning"}],
                "ai": {"provider": "deepseek_v4_pro"},
            }
        )
        response_payload = {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "summary": "no safe field mapping",
                                "reasons": ["headers do not map"],
                                "missing_information": ["field mapping"],
                                "suggested_target_tables": ["goods"],
                                "suggested_field_mappings": [],
                                "next_steps": ["add mapping"],
                            }
                        )
                    }
                }
            ]
        }
        captured = {}

        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def read(self):
                return json.dumps(response_payload).encode("utf-8")

        def fake_urlopen(request, timeout):
            captured["url"] = request.full_url
            captured["body"] = json.loads(request.data.decode("utf-8"))
            return FakeResponse()

        with patch.dict(os.environ, {"BASEAI_API_KEY": "unit-key"}, clear=True):
            with patch("urllib.request.urlopen", side_effect=fake_urlopen):
                review = call_draft_diagnostics_ai(manifest, {"patch": {"operation_count": 0}})

        self.assertEqual(review["suggested_target_tables"], ["goods"])
        self.assertEqual(captured["url"], "https://baseai.rivergame.net/v1/chat/completions")
        self.assertIn("zero operations", captured["body"]["messages"][0]["content"])


if __name__ == "__main__":
    unittest.main()
