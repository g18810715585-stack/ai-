from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "fixtures"


def create_planning() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "礼包规划"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "六月礼包"
    sheet["A2"] = "礼包ID"
    sheet["B2"] = "礼包名称"
    sheet["C2"] = "价格ID"
    sheet["D2"] = "限购次数"
    sheet["E2"] = "开始时间"
    sheet["F2"] = "结束时间"
    sheet["A3"] = 1001
    sheet["B3"] = "每日礼包"
    sheet["C3"] = 68
    sheet["D3"] = None
    sheet["E3"] = "2026-06-01 05:00:00"
    sheet["F3"] = "2026-06-08 04:59:59"
    sheet["A4"] = 1002
    sheet["B4"] = "进阶礼包"
    sheet["C4"] = 128
    sheet["D4"] = 1
    sheet["E4"] = "2026-06-01 05:00:00"
    sheet["F4"] = "2026-06-08 04:59:59"
    sheet["B4"].fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    sheet["F4"].comment = Comment("示例备注", "ai-meta-agent")
    sheet.column_dimensions["G"].hidden = True
    sheet["G2"] = "隐藏ID"
    sheet["G3"] = "hidden-1001"
    workbook.save(FIXTURES / "sample-planning.xlsx")


def create_config() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "shop_pack_config"
    headers = [
        "pack_id",
        "name",
        "price_id",
        "limit_type",
        "limit_count",
        "start_time",
        "end_time",
        "creator",
        "create_time",
        "internal_note",
    ]
    for col, header in enumerate(headers, start=1):
        sheet.cell(1, col).value = header
    values = [
        1001,
        "旧每日礼包",
        30,
        "daily",
        1,
        "2026-05-01 05:00:00",
        "2026-05-08 04:59:59",
        "gaoyang",
        "2026-05-01",
        "keep me",
    ]
    for col, value in enumerate(values, start=1):
        sheet.cell(2, col).value = value
    reward = workbook.create_sheet("reward_item_config")
    for col, header in enumerate(["reward_group_id", "item_id", "item_count", "item_order"], start=1):
        reward.cell(1, col).value = header
    reward.append([1001, 3001, 10, 1])
    workbook.save(FIXTURES / "sample-config.xlsx")


def main() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)
    create_planning()
    create_config()
    print(f"created fixtures in {FIXTURES}")


if __name__ == "__main__":
    main()
